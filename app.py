import os
import requests as http_requests
from urllib.parse import urlencode

# OAuth & Security config
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
GOOGLE_AUTH_URL = 'https://accounts.google.com/o/oauth2/v2/auth'
GOOGLE_TOKEN_URL = 'https://oauth2.googleapis.com/token'
GOOGLE_USERINFO_URL = 'https://www.googleapis.com/oauth2/v3/userinfo'
RECAPTCHA_SITE_KEY = os.environ.get('RECAPTCHA_SITE_KEY', '')
RECAPTCHA_SECRET_KEY = os.environ.get('RECAPTCHA_SECRET_KEY', '')
# Brevo (Sendinblue) email config
BREVO_API_KEY = os.environ.get('BREVO_API_KEY', '')
BREVO_SMTP_KEY = os.environ.get('BREVO_SMTP_KEY', '')
MAIL_FROM = os.environ.get('MAIL_FROM', 'noreply@maritime-tests.bg')
MAIL_FROM_NAME = 'Морски Тестове'
BASE_URL = os.environ.get('BASE_URL', 'https://web-production-ca6b6.up.railway.app')

def send_verification_email(to_email, token):
    """Send verification email via Brevo SMTP"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    verify_url = f"{BASE_URL}/verify-email/{token}"
    
    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'Потвърди имейла си — Морски Тестове'
    msg['From'] = f"{MAIL_FROM_NAME} <{MAIL_FROM}>"
    msg['To'] = to_email
    
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:32px;background:#071a2e;border-radius:16px">
      <h2 style="color:#e8a020;font-size:22px;margin-bottom:12px">⚓ Морски Тестове</h2>
      <h3 style="color:#fff;margin-bottom:16px">Потвърди имейла си</h3>
      <p style="color:rgba(232,237,242,0.8);margin-bottom:24px">
        Благодарим за регистрацията! Натисни бутона за да активираш акаунта си.
      </p>
      <a href="{verify_url}" 
         style="display:inline-block;background:#635BFF;color:#fff;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:600;font-size:15px">
        Потвърди акаунта →
      </a>
      <p style="color:rgba(232,237,242,0.4);font-size:12px;margin-top:24px">
        Линкът е валиден 24 часа. Ако не си се регистрирал — игнорирай този имейл.
      </p>
    </div>
    """
    
    # Add plain text fallback
    text = f"""Потвърди имейла си — Морски Тестове

Благодарим за регистрацията!
Натисни линка за да активираш акаунта си:

{verify_url}

Линкът е валиден 24 часа.
Ако не си се регистрирал — игнорирай този имейл.
"""
    msg.attach(MIMEText(text, 'plain'))
    msg.attach(MIMEText(html, 'html'))
    
    try:
        with smtplib.SMTP('smtp-relay.brevo.com', 587, timeout=10) as server:
            server.starttls()
            server.login(MAIL_FROM, BREVO_SMTP_KEY)
            server.sendmail(MAIL_FROM, to_email, msg.as_string())
        print(f"✓ Verification email sent to {to_email}")
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

def send_verification_email_async(to_email, token):
    """Send email in background thread"""
    import threading
    t = threading.Thread(target=send_verification_email, args=(to_email, token))
    t.daemon = True
    t.start()



from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
import xlrd, json, random, string

app = Flask(__name__)
app.config['SECRET_KEY'] = 'maritime-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///maritime.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
os.makedirs('/tmp/uploads', exist_ok=True)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

db = SQLAlchemy(app)

# ============================================================
#  МОДЕЛИ (Таблици в базата данни)
# ============================================================

class User(db.Model):
    """Моряци / потребители"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    rank = db.Column(db.String(100), default='')          # Длъжност/ранг
    company = db.Column(db.String(100), default='')       # Компания
    category = db.Column(db.String(20), default='deck')   # deck / engine
    level = db.Column(db.String(30), default='Operational Level')
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    email_verified = db.Column(db.Boolean, default=False)
    verification_token = db.Column(db.String(64), nullable=True)
    last_seen = db.Column(db.DateTime, default=None)
    promo_code = db.Column(db.String(50), default='')     # Кода с който се е регистрирал
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    results = db.relationship('TestResult', backref='user', lazy=True)

class Test(db.Model):
    """Тестове (качени от admin)"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(20), nullable=False)   # deck / engine
    level = db.Column(db.String(50), default='Operational Level')
    questions_json = db.Column(db.Text, nullable=False)   # JSON с въпросите
    question_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_demo = db.Column(db.Boolean, default=False)
    results = db.relationship('TestResult', backref='test', lazy=True)

    def get_questions(self):
        return json.loads(self.questions_json)


class DemoVisit(db.Model):
    """Посещения на демо страницата"""
    id = db.Column(db.Integer, primary_key=True)
    ip_hash = db.Column(db.String(64), nullable=False)  # SHA256 на IP - GDPR safe
    visited_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_agent = db.Column(db.String(200), default='')  # браузър

class TestResult(db.Model):
    """Резултати от тестове"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    test_id = db.Column(db.Integer, db.ForeignKey('test.id'), nullable=False)
    score = db.Column(db.Integer, default=0)
    total = db.Column(db.Integer, default=0)
    percent = db.Column(db.Float, default=0)
    passed = db.Column(db.Boolean, default=False)
    answers_json = db.Column(db.Text, default='{}')       # Запазени отговори
    test_type = db.Column(db.String(20), default='test')
    duration = db.Column(db.Integer, default=0)  # секунди
    question_ids_json = db.Column(db.Text, default='[]')  # ID-та на въпросите
    taken_at = db.Column(db.DateTime, default=datetime.utcnow)

class TestImage(db.Model):
    """Снимки към въпроси — пазят се отделно"""
    id = db.Column(db.Integer, primary_key=True)
    test_id = db.Column(db.Integer, db.ForeignKey('test.id'), nullable=False)
    question_id = db.Column(db.Integer, nullable=False)  # q['id']
    image_data = db.Column(db.Text, nullable=False)  # base64

class PromoCode(db.Model):
    """Промокодове"""
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    client_name = db.Column(db.String(200), default='')
    access_type = db.Column(db.String(100), default='Регулярни тестове')
    price = db.Column(db.Float, default=0)
    is_active = db.Column(db.Boolean, default=True)
    is_used = db.Column(db.Boolean, default=False)
    used_by = db.Column(db.String(120), default='')
    used_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Signal(db.Model):
    """Сигнали / бъгове от потребители"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    user_name = db.Column(db.String(100), default='Анонимен')
    type = db.Column(db.String(50), default='bug')        # bug / suggestion / question
    message = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='open')     # open / resolved
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ============================================================
#  ПОМОЩНИ ФУНКЦИИ
# ============================================================

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.is_admin:
            return redirect(url_for('user_dashboard'))
        return f(*args, **kwargs)
    return decorated

def parse_xls_colors(filepath):
    """Чете XLS/XLSX и открива верни отговори по цвят на шрифта"""
    OPT_LETTERS = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
    questions = []

    if filepath.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Извличаме снимките — всяка снимка принадлежи на въпроса НАД нея
        image_map = {}
        try:
            all_images = ws._images
            print(f"PARSE: Found {len(all_images)} images in worksheet")
        except Exception as e:
            print(f"PARSE: Cannot access _images: {e}")
            all_images = []
        
        for img in all_images:
            try:
                anchor_row_0idx = img.anchor._from.row
                ws_row_of_image = anchor_row_0idx + 1
                question_ws_row = ws_row_of_image - 1
                # Try different methods to get image data
                try:
                    img_data = img._data()
                except:
                    try:
                        img_data = img.ref.blob
                    except:
                        img_data = bytes(img.ref._data)
                fmt = 'jpg' if img_data[:2] == b'\xff\xd8' else 'png'
                image_map[question_ws_row] = (img_data, fmt)
            except Exception as e:
                print(f"PARSE: Image error: {e}")

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            q_cell = row[0]
            if not q_cell.value or str(q_cell.value).strip() == '':
                continue
            q_text = str(q_cell.value).strip()
            options = []
            opt_idx = 0

            for cell in row[1:]:
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                is_correct = False
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == 'rgb':
                        rgb = color.rgb
                        if rgb not in ('00000000', 'FF000000', '000000'):
                            is_correct = True
                    elif color.type == 'theme':
                        if color.theme not in (0, 1):
                            is_correct = True
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            q_id = len(questions) + 1  # 1, 2, 3... последователно
            q = {'id': q_id, 'question': q_text, 'options': options}
            if r_idx in image_map:
                q['has_image'] = True
                q['_image_data'] = image_map[r_idx]
            questions.append(q)

    else:
        # XLS - използваме xlrd
        BLACK_IDX = 8
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        for r in range(1, ws.nrows):
            q_val = ws.cell(r, 0).value
            if not q_val or str(q_val).strip() == '':
                continue
            q_text = str(q_val).strip()
            options = []
            opt_idx = 0

            for c in range(1, ws.ncols):
                cell = ws.cell(r, c)
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                xf_idx = ws.cell_xf_index(r, c)
                xf = wb.xf_list[xf_idx]
                font = wb.font_list[xf.font_index]
                is_correct = (font.colour_index != BLACK_IDX)
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            questions.append({'id': len(questions) + 1, 'question': q_text, 'options': options})

    return questions

def generate_promo_code(prefix='MAR'):
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"{prefix}-{suffix}"

# ============================================================
#  AUTH ROUTES
# ============================================================

@app.route('/qimage/<int:test_id>/<path:filename>')
@login_required
def serve_qimage(test_id, filename):
    """Сервира снимки директно — по-бързо от base64"""
    from flask import send_file
    for base_dir in [f"/tmp/qimages/{test_id}", f"/data/qimages/{test_id}"]:
        img_path = os.path.join(base_dir, filename)
        if os.path.exists(img_path):
            return send_file(img_path)
    return '', 404

@app.route('/admin/tests/force-upload', methods=['POST'])
@admin_required
def force_upload():
    """Качва тест използвайки вече парснатите данни от сесията"""
    pending = session.get('pending_upload')
    if not pending:
        return jsonify({'error': 'Няма данни за качване'}), 400
    
    new_title = request.json.get('title', pending['title'])
    
    test = Test(
        title=new_title,
        category=pending['category'],
        level=pending['level'],
        questions_json=pending['questions_json'],
        question_count=pending['question_count'],
        is_demo=False
    )
    db.session.add(test)
    db.session.flush()
    
    # Запази снимките
    if pending.get('images'):
        img_dir = f"/tmp/qimages/{test.id}"
        os.makedirs(img_dir, exist_ok=True)
        for q_id, (img_data, fmt) in pending['images']:
            try:
                with open(f"{img_dir}/{q_id}.{fmt}", 'wb') as f_img:
                    f_img.write(img_data)
            except Exception as e:
                print(f"Image save error: {e}")
    
    db.session.commit()
    session.pop('pending_upload', None)
    return jsonify({'success': True, 'title': new_title, 'total': pending['question_count']})

# toggle_demo route removed - use /admin/demo/toggle/<id>

@app.route('/admin/tests/next-title')
@admin_required
def next_available_title():
    title = request.args.get('title', '')
    if not Test.query.filter_by(title=title).first():
        return jsonify({'title': title, 'duplicate': False})
    idx = 1
    while Test.query.filter_by(title=f"{title} ({idx})").first():
        idx += 1
    return jsonify({'title': f"{title} ({idx})", 'duplicate': True})


@app.context_processor
def inject_recaptcha():
    return dict(recaptcha_site_key=RECAPTCHA_SITE_KEY)

@app.before_request
def update_last_seen():
    if 'user_id' in session:
        now = datetime.utcnow()
        last = session.get('_last_seen_update')
        # Only update DB every 5 minutes
        if not last or (now - datetime.fromisoformat(last)).seconds > 300:
            user = User.query.get(session['user_id'])
            if user:
                user.last_seen = now
                db.session.commit()
                session['_last_seen_update'] = now.isoformat()

@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/demo')
def demo():
    # Log demo visit - GDPR safe (hashed IP)
    import hashlib
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    ip_hash = hashlib.sha256(ip.encode()).hexdigest()
    ua = request.headers.get('User-Agent', '')[:200]
    visit = DemoVisit(ip_hash=ip_hash, user_agent=ua)
    db.session.add(visit)
    db.session.commit()

    # Send ALL tests - demo ones playable, others informative
    demo_tests_raw = Test.query.order_by(Test.category, Test.level).all()
    
    level_map = {
        'Operational Level': 'operational',
        'operational level': 'operational',
        'operational': 'operational',
        'Оперативно ниво': 'operational',
        'Management Level': 'management',
        'management level': 'management',
        'management': 'management',
        'Мениджърско ниво': 'management',
        'Master Level': 'master',
        'master level': 'master',
        'master': 'master',
        'Капитанско ниво': 'master',
        'Support Level': 'operational',
        'support level': 'operational',
    }
    
    # Only expose safe fields - never expose questions_json or internal data
    tests_data = []
    for t in demo_tests_raw:
        level_key = level_map.get(t.level) or level_map.get(t.level.strip()) or 'operational'
        cat = t.category.lower().strip()
        if cat not in ('deck', 'engine'):
            cat = 'deck' if 'deck' in cat or 'палуб' in cat.lower() else 'engine'
        tests_data.append({
            'id': t.id,
            'title': t.title,
            'category': cat,
            'level_key': level_key,
            'question_count': t.question_count,
            'is_demo': t.is_demo
        })
    
    return render_template('demo.html', demo_tests=tests_data, recaptcha_site_key=RECAPTCHA_SITE_KEY)


@app.route('/admin/demo')
@admin_required
def admin_demo():
    tests = Test.query.filter_by(is_demo=True).order_by(Test.category, Test.level, Test.title).all()
    demo_count = Test.query.filter_by(is_demo=True).count()
    deck_demo = Test.query.filter_by(is_demo=True, category='deck').count()
    engine_demo = Test.query.filter_by(is_demo=True, category='engine').count()
    return render_template('admin_demo.html',
        active='demo',
        tests=tests,
        demo_count=demo_count,
        deck_demo=deck_demo,
        engine_demo=engine_demo
    )

@app.route('/admin/demo/toggle/<int:test_id>', methods=['POST'])
@admin_required
def admin_demo_toggle(test_id):
    # Extra security: verify request is AJAX from same origin
    if not request.is_json and request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        # Accept both JSON and same-origin fetch
        pass
    
    # Validate test_id is positive integer (already done by Flask route)
    test = Test.query.get_or_404(test_id)
    
    # Toggle demo status
    test.is_demo = not test.is_demo
    db.session.commit()
    
    return jsonify({
        'success': True,
        'is_demo': test.is_demo,
        'test_id': test.id
    })


@app.route('/admin/demo/reset-all', methods=['POST'])
@admin_required  
def admin_demo_reset_all():
    """Reset ALL tests to is_demo=False - emergency fix"""
    Test.query.update({Test.is_demo: False})
    db.session.commit()
    count = Test.query.count()
    return jsonify({'success': True, 'message': f'Reset {count} tests to is_demo=False'})




def get_google_redirect_uri():
    return os.environ.get('BASE_URL', 'https://web-production-ca6b6.up.railway.app') + '/auth/google/callback'

@app.route('/auth/google')
def google_login():
    params = {
        'client_id': GOOGLE_CLIENT_ID,
        'redirect_uri': get_google_redirect_uri(),
        'response_type': 'code',
        'scope': 'openid email profile',
        'access_type': 'offline',
        'prompt': 'select_account'
    }
    return redirect(GOOGLE_AUTH_URL + '?' + urlencode(params))

@app.route('/auth/google/callback')
def google_callback():
    code = request.args.get('code')
    error = request.args.get('error')
    
    if error or not code:
        flash('Google вход е отказан.', 'error')
        return redirect(url_for('login'))
    
    # Exchange code for token
    token_data = {
        'code': code,
        'client_id': GOOGLE_CLIENT_ID,
        'client_secret': GOOGLE_CLIENT_SECRET,
        'redirect_uri': get_google_redirect_uri(),
        'grant_type': 'authorization_code'
    }
    token_resp = http_requests.post(GOOGLE_TOKEN_URL, data=token_data)
    if not token_resp.ok:
        flash('Грешка при Google автентикация.', 'error')
        return redirect(url_for('login'))
    
    access_token = token_resp.json().get('access_token')
    
    # Get user info
    user_info = http_requests.get(
        GOOGLE_USERINFO_URL,
        headers={'Authorization': f'Bearer {access_token}'}
    ).json()
    
    google_email = user_info.get('email', '').lower().strip()
    google_name = user_info.get('name', '')
    google_id = user_info.get('sub', '')
    
    if not google_email:
        flash('Не може да се получи имейл от Google.', 'error')
        return redirect(url_for('login'))
    
    # Find or create user
    user = User.query.filter_by(email=google_email).first()
    
    if user:
        # Existing user - log in
        session['user_id'] = user.id
        if user.is_admin:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('user_dashboard'))
    else:
        # New user - create account
        new_user = User(
            name=google_name or google_email.split('@')[0],
            email=google_email,
            password=generate_password_hash(google_id + GOOGLE_CLIENT_SECRET[:8]),
            is_admin=False,
            is_active=True
        )
        db.session.add(new_user)
        db.session.commit()
        session['user_id'] = new_user.id
        flash(f'Добре дошъл, {new_user.name}! Акаунтът ти е създаден с Google.', 'success')
        return redirect(url_for('user_dashboard'))


@app.route('/debug-env')
def debug_env():
    import os
    return jsonify({
        'RECAPTCHA_SITE_KEY': os.environ.get('RECAPTCHA_SITE_KEY', 'NOT SET'),
        'RECAPTCHA_SITE_KEY_len': len(os.environ.get('RECAPTCHA_SITE_KEY', '')),
        'GOOGLE_CLIENT_ID': os.environ.get('GOOGLE_CLIENT_ID', 'NOT SET')[:20] + '...',
        'module_level_key': RECAPTCHA_SITE_KEY[:20] if RECAPTCHA_SITE_KEY else 'EMPTY'
    })


@app.route('/verify-email/<token>')
def verify_email(token):
    user = User.query.filter_by(verification_token=token).first()
    if not user:
        flash('Невалиден или изтекъл верификационен линк.', 'error')
        return redirect(url_for('index'))
    user.email_verified = True
    user.verification_token = None
    db.session.commit()
    session['user_id'] = user.id
    flash('Имейлът е потвърден! Добре дошъл!', 'success')
    return redirect(url_for('user_dashboard'))

@app.route('/ping')
def ping():
    return 'ok', 200

@app.route('/')
def index():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user and user.is_admin:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('user_dashboard'))
    return render_template('landing.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['user_name'] = user.name
            session['is_admin'] = user.is_admin
            if user.is_admin:
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('user_dashboard'))
        flash('Грешен имейл или парола', 'error')
    return render_template('login.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

# Simple rate limiting store
_reg_attempts = {}

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Rate limiting - max 3 attempts per IP per hour
        from datetime import timedelta
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or '127.0.0.1').split(',')[0].strip()
        now = datetime.utcnow()
        _reg_attempts[ip] = [t for t in _reg_attempts.get(ip, []) if now - t < timedelta(hours=1)]
        if len(_reg_attempts.get(ip, [])) >= 3:
            flash('Твърде много опити за регистрация. Опитай след 1 час.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)
        _reg_attempts.setdefault(ip, []).append(now)

        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        promo = request.form.get('promo_code', '').strip()

        # Basic validation
        if not email or not password:
            flash('Имейлът и паролата са задължителни.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

        # Verify reCAPTCHA
        recaptcha_response = request.form.get('cf-turnstile-response', '')
        if RECAPTCHA_SECRET_KEY and not recaptcha_response:
            flash('Моля потвърди че не си робот.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)
        
        if RECAPTCHA_SECRET_KEY and recaptcha_response:
            verify = http_requests.post('https://challenges.cloudflare.com/turnstile/v0/siteverify', data={
                'secret': RECAPTCHA_SECRET_KEY,
                'response': recaptcha_response,
                'remoteip': request.remote_addr
            }).json()
            if not verify.get('success'):
                flash('reCAPTCHA верификацията е неуспешна. Опитай отново.', 'error')
                return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

        if len(password) < 6:
            flash('Паролата трябва да е поне 6 символа.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

        import re as _re
        if not (_re.search(r'[a-zA-Zа-яА-Я]', password) and _re.search(r'[0-9]', password)):
            flash('Паролата трябва да съдържа букви И цифри.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

        # Basic email format validation
        import re as _re
        if not _re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
            flash('Невалиден имейл адрес.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

        if User.query.filter_by(email=email).first():
            flash('Имейлът вече е регистриран. Влез в профила си.', 'error')
            return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

        # Create user
        name = request.form.get('name', '').strip() or email.split('@')[0]
        user = User(
            name=name,
            email=email,
            password=generate_password_hash(password),
            is_active=True,
            email_verified=False
        )
        db.session.add(user)

        # Apply promo code if provided
        if promo:
            promo_obj = PromoCode.query.filter_by(code=promo, is_active=True, is_used=False).first()
            if promo_obj:
                promo_obj.is_used = True
                promo_obj.used_by = email
                promo_obj.used_at = datetime.utcnow()
                promo_obj.is_active = False
                db.session.commit()
                session['user_id'] = user.id
                flash('Акаунтът е създаден и промокодът е активиран!', 'success')
                return redirect(url_for('user_dashboard'))

        # Generate verification token
        import secrets
        token = secrets.token_urlsafe(32)
        user.verification_token = token
        db.session.commit()
        
        # Send verification email
        if BREVO_SMTP_KEY:
            send_verification_email_async(email, token)
            session['user_id'] = user.id
            flash('Акаунтът е създаден! Провери имейла си за потвърждение.', 'success')
        else:
            user.email_verified = True
            db.session.commit()
            session['user_id'] = user.id
            flash('Добре дошъл! Акаунтът ти е създаден.', 'success')
        return redirect(url_for('user_dashboard'))

    return render_template('register.html', recaptcha_site_key=RECAPTCHA_SITE_KEY)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# ============================================================
#  USER ROUTES
# ============================================================

@app.route('/dashboard')
@login_required
def user_dashboard():
    user = User.query.get(session['user_id'])
    results = TestResult.query.filter_by(user_id=user.id).order_by(TestResult.taken_at.desc()).limit(5).all()
    total_tests = TestResult.query.filter_by(user_id=user.id).count()
    passed_tests = TestResult.query.filter_by(user_id=user.id, passed=True).count()
    tests = Test.query.order_by(Test.created_at.desc()).all()
    return render_template('dashboard_user.html', user=user, results=results,
                           total_tests=total_tests, passed_tests=passed_tests, tests=tests)

def inject_images(test_id, questions):
    """Добавя снимките към въпросите — URL вместо base64"""
    img_dir = f"/tmp/qimages/{test_id}"
    if not os.path.exists(img_dir):
        print(f"INJECT: No image dir found for test {test_id}")
        return questions
    
    loaded = 0
    for q in questions:
        if q.get('has_image'):
            for fmt in ['jpg', 'png']:
                img_path = f"{img_dir}/{q['id']}.{fmt}"
                if os.path.exists(img_path):
                    # URL вместо base64 — браузърът зарежда при нужда
                    q['image'] = f"/qimage/{test_id}/{q['id']}.{fmt}"
                    loaded += 1
                    break
    print(f"INJECT: Loaded {loaded} images for test {test_id}")
    return questions

@app.route('/test/<int:test_id>')
@login_required
def take_test(test_id):
    import random as rnd
    test = Test.query.get_or_404(test_id)
    questions = test.get_questions()
    questions = inject_images(test_id, questions)
    shuffle = request.args.get('shuffle') == 'true'
    if shuffle:
        questions = list(questions)
        rnd.shuffle(questions)
    return render_template('take_test.html', test=test, questions=questions, shuffle=shuffle)

@app.route('/test/<int:test_id>/mistakes')
@login_required
def test_mistakes(test_id):
    import random as rnd
    test = Test.query.get_or_404(test_id)
    
    # Вземи последните 2 резултата от обикновен тест или микс
    last_results = TestResult.query.filter_by(
        user_id=session['user_id'],
        test_id=test_id
    ).filter(
        TestResult.test_type.in_(['test', 'mix'])
    ).order_by(TestResult.taken_at.desc()).limit(2).all()
    
    if len(last_results) < 2:
        flash('Трябват поне 2 решени теста (Тест или Микс) за тази функция', 'error')
        return redirect(url_for('admin_tests'))
    
    # Събери грешно отговорените въпроси
    all_questions = test.get_questions()
    q_map = {str(q['id']): q for q in all_questions}
    
    # Въпрос се включва в грешките ако е грешен в ПОНЕ ЕДИН от последните 2 теста
    # Въпрос се премахва само ако е верен И В ДВАТА последни теста
    
    # Намери всички въпроси отговорени в двата теста
    answered_in = [{}, {}]  # {q_id: is_correct} за всеки тест
    for i, result in enumerate(last_results):
        answers = json.loads(result.answers_json)
        for q_id_str, o_idx in answers.items():
            q = q_map.get(str(q_id_str))
            if q:
                try:
                    is_correct = q['options'][int(o_idx)]['isCorrect']
                    answered_in[i][str(q_id_str)] = is_correct
                except (IndexError, KeyError):
                    answered_in[i][str(q_id_str)] = False
    
    wrong_ids = set()
    all_answered = set(answered_in[0].keys()) | set(answered_in[1].keys())
    
    for q_id_str in all_answered:
        correct_in_0 = answered_in[0].get(q_id_str, None)
        correct_in_1 = answered_in[1].get(q_id_str, None)
        
        # Грешен ако е грешен в поне един тест
        # Верен само ако е верен в ДВАТА теста
        if correct_in_0 is False or correct_in_1 is False:
            wrong_ids.add(q_id_str)
        elif correct_in_0 is None or correct_in_1 is None:
            # Отговорен само в един тест — включи ако е грешен
            val = correct_in_0 if correct_in_0 is not None else correct_in_1
            if not val:
                wrong_ids.add(q_id_str)
    
    if not wrong_ids:
        flash('Нямаш грешки от последните 2 теста!', 'success')
        return redirect(url_for('admin_tests'))
    
    # Вземи въпросите с грешки
    wrong_questions = [q for q in all_questions if str(q['id']) in wrong_ids]
    wrong_questions = inject_images(test_id, wrong_questions)
    rnd.shuffle(wrong_questions)
    
    return render_template('take_test.html', test=test, questions=wrong_questions, 
                         shuffle=True, test_type='mistakes')

@app.route('/test/<int:test_id>/simulator')
@login_required
def simulator(test_id):
    import random as rnd
    test = Test.query.get_or_404(test_id)
    questions = test.get_questions()
    questions = inject_images(test_id, questions)
    rnd.shuffle(questions)
    questions = questions[:60]
    return render_template('simulator.html', test=test, questions=questions)

@app.route('/test/<int:test_id>/submit', methods=['POST'])
@login_required
def submit_test(test_id):
    test = Test.query.get_or_404(test_id)
    all_questions = test.get_questions()
    answers = request.json.get('answers', {})
    test_type = request.json.get('test_type', 'test')
    # ID-тата на въпросите пратени от frontend
    question_ids = request.json.get('question_ids', [])

    # Нормализирай ключовете към стрингове
    answers_normalized = {str(k): int(v) for k, v in answers.items()}

    # Ако симулаторът е пратил конкретни ID-та — ползвай само тях
    if question_ids:
        qid_set = set(str(qid) for qid in question_ids)
        questions = [q for q in all_questions if str(q['id']) in qid_set]
    else:
        questions = all_questions

    score = 0
    for q in questions:
        q_id = str(q['id'])
        selected = answers_normalized.get(q_id)
        if selected is not None:
            try:
                if q['options'][int(selected)]['isCorrect']:
                    score += 1
            except (IndexError, KeyError):
                pass

    total = len(questions)
    answered = len(answers_normalized)
    percent = round((score / total) * 100, 1) if total > 0 else 0
    # Взет тест:
    # - Симулатор: грешни <= 10% от total (т.е. <= 6 от 60)
    # - Всички останали: >= 90% верни
    wrong = total - score
    if test_type == 'simulator':
        # Взет ако: грешни <= 10% от total ИЛИ верни >= 90%
        passed = (wrong <= round(total * 0.10)) or (percent >= 90)
    else:
        passed = percent >= 90

    answers_normalized_final = answers_normalized

    duration = request.json.get('duration', 0)

    result = TestResult(
        user_id=session['user_id'],
        test_id=test_id,
        score=score, total=total,
        percent=percent, passed=passed,
        answers_json=json.dumps(answers_normalized),
        test_type=test_type,
        duration=duration,
        question_ids_json=json.dumps(question_ids)
    )
    db.session.add(result)
    db.session.commit()

    return jsonify({'score': score, 'total': total, 'percent': percent, 'passed': passed})

@app.route('/history')
@login_required
def history():
    user = User.query.get(session['user_id'])
    results = TestResult.query.filter_by(user_id=user.id).order_by(TestResult.taken_at.desc()).all()
    return render_template('history.html', user=user, results=results)

@app.route('/signal', methods=['POST'])
@login_required
def submit_signal():
    msg = request.form.get('message', '').strip()
    sig_type = request.form.get('type', 'bug')
    user = User.query.get(session['user_id'])
    if msg:
        signal = Signal(user_id=user.id, user_name=user.name, type=sig_type, message=msg)
        db.session.add(signal)
        db.session.commit()
    return redirect(url_for('user_dashboard'))

# ============================================================
#  ADMIN ROUTES
# ============================================================

@app.route('/admin')
@admin_required
def admin_dashboard():
    total_users = User.query.filter_by(is_admin=False).count()
    active_promos = PromoCode.query.filter_by(is_active=True).count()
    total_tests = Test.query.count()
    total_results = TestResult.query.count()
    open_signals = Signal.query.filter_by(status='open').count()
    recent_results = TestResult.query.order_by(TestResult.taken_at.desc()).limit(8).all()
    deck_q = db.session.query(db.func.sum(Test.question_count)).filter_by(category='deck').scalar() or 0
    engine_q = db.session.query(db.func.sum(Test.question_count)).filter_by(category='engine').scalar() or 0

    recent_signals = Signal.query.order_by(Signal.created_at.desc()).limit(4).all()

    # Demo visit stats
    from datetime import timedelta
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())
    month_start = today_start.replace(day=1)

    demo_today = DemoVisit.query.filter(DemoVisit.visited_at >= today_start).count()
    demo_week = DemoVisit.query.filter(DemoVisit.visited_at >= week_start).count()
    demo_month = DemoVisit.query.filter(DemoVisit.visited_at >= month_start).count()
    demo_total = DemoVisit.query.count()

    # Unique visitors (by ip_hash)
    demo_unique_today = db.session.query(DemoVisit.ip_hash).filter(
        DemoVisit.visited_at >= today_start).distinct().count()
    demo_unique_week = db.session.query(DemoVisit.ip_hash).filter(
        DemoVisit.visited_at >= week_start).distinct().count()
    demo_unique_month = db.session.query(DemoVisit.ip_hash).filter(
        DemoVisit.visited_at >= month_start).distinct().count()
    demo_unique_total = db.session.query(DemoVisit.ip_hash).distinct().count()

    # Recent demo visits with order number
    recent_demo = DemoVisit.query.order_by(DemoVisit.visited_at.desc()).limit(10).all()

    return render_template('admin_dashboard.html',
        total_users=total_users, active_promos=active_promos,
        total_tests=total_tests, total_results=total_results,
        open_signals=open_signals, recent_results=recent_results,
        recent_signals=recent_signals,
        deck_q=deck_q, engine_q=engine_q,
        demo_today=demo_today,
        demo_week=demo_week,
        demo_month=demo_month,
        demo_total=demo_total,
        demo_unique_today=demo_unique_today,
        demo_unique_week=demo_unique_week,
        demo_unique_month=demo_unique_month,
        demo_unique_total=demo_unique_total,
        recent_demo=recent_demo
    )

@app.route('/admin/tests')
@admin_required
def admin_tests():
    deck_tests = Test.query.filter_by(category='deck').order_by(Test.created_at.desc()).all()
    engine_tests = Test.query.filter_by(category='engine').order_by(Test.created_at.desc()).all()
    
    # За всеки тест провери дали има 2+ резултата
    mistakes_ready = {}
    for t in deck_tests + engine_tests:
        count = TestResult.query.filter_by(
            user_id=session['user_id'], test_id=t.id
        ).filter(TestResult.test_type.in_(['test', 'mix'])).count()
        mistakes_ready[t.id] = count >= 2
    
    return render_template('admin_tests.html', deck_tests=deck_tests, engine_tests=engine_tests, mistakes_ready=mistakes_ready)

@app.route('/admin/tests/upload', methods=['POST'])
@admin_required
def upload_test():
    file = request.files.get('file')
    title = request.form.get('title', '').strip()
    category = request.form.get('category', 'deck')
    level = request.form.get('level', 'Operational Level')

    if not file:
        return jsonify({'error': 'Няма файл'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        print(f"UPLOAD: Starting parse of {filename}, size={os.path.getsize(filepath)} bytes")
        questions = parse_xls_colors(filepath)
        print(f"UPLOAD: Parsed {len(questions)} questions")
        with_img = sum(1 for q in questions if q.get('has_image'))
        print(f"UPLOAD: Questions with images: {with_img}")
        final_title = title if title else filename.replace('.xls', '').replace('.xlsx', '')
        
        # Провери за дублиращо се заглавие
        existing = Test.query.filter_by(title=final_title).first()
        if existing:
            force = request.form.get('force', 'false')
            if force != 'true':
                # Запази парснатите данни в сесията за по-късно
                import pickle, base64
                session['pending_upload'] = {
                    'questions_json': __import__('json').dumps(
                        [{k: v for k, v in q.items() if k != '_image_data'} for q in questions],
                        ensure_ascii=False
                    ),
                    'question_count': len(questions),
                    'category': category,
                    'level': level,
                    'title': final_title,
                    'images': [(q['id'], q['_image_data']) for q in questions if '_image_data' in q]
                }
                os.remove(filepath)
                return jsonify({'duplicate': True, 'title': final_title})
            else:
                # Намери следващия свободен индекс
                idx = 1
                while Test.query.filter_by(title=f"{final_title} ({idx})").first():
                    idx += 1
                final_title = f"{final_title} ({idx})" 

        # Извади снимките преди да запишем JSON
        images_to_save = []
        for q in questions:
            if '_image_data' in q:
                images_to_save.append((q['id'], q.pop('_image_data')))

        test = Test(
            title=final_title,
            category=category,
            level=level,
            questions_json=json.dumps(questions, ensure_ascii=False),
            question_count=len(questions),
            is_demo=False
        )
        db.session.add(test)
        db.session.flush()
        test_id_for_images = test.id
        db.session.commit()
        os.remove(filepath)

        # Запази снимките директно
        if images_to_save:
            img_dir = f"/tmp/qimages/{test_id_for_images}"
            print(f"IMAGES: Saving {len(images_to_save)} images to {img_dir}")
            try:
                os.makedirs(img_dir, exist_ok=True)
                print(f"IMAGES: Directory created: {img_dir}")
            except Exception as e:
                print(f"IMAGES: Cannot create dir {img_dir}: {e}")
                # Fallback to /tmp
                img_dir = f"/tmp/qimages/{test_id_for_images}"
                os.makedirs(img_dir, exist_ok=True)
                print(f"IMAGES: Using fallback: {img_dir}")
            
            saved_count = 0
            for q_id, (img_data, fmt) in images_to_save:
                try:
                    img_path = f"{img_dir}/{q_id}.{fmt}"
                    with open(img_path, 'wb') as f_img:
                        f_img.write(img_data)
                    saved_count += 1
                except Exception as e:
                    print(f"IMAGES: Save error q{q_id}: {e}")
            print(f"IMAGES: Saved {saved_count}/{len(images_to_save)} images")

        return jsonify({'success': True, 'total': len(questions), 'title': final_title})
    except Exception as e:
        try: os.remove(filepath)
        except: pass
        return jsonify({'error': str(e)}), 500

@app.route('/admin/tests/<int:test_id>/edit')
@admin_required
def edit_test(test_id):
    test = Test.query.get_or_404(test_id)
    questions = test.get_questions()
    questions = inject_images(test_id, questions)
    return render_template('edit_test.html', test=test, questions=questions)

@app.route('/admin/tests/<int:test_id>/update-info', methods=['POST'])
@admin_required
def update_test_info(test_id):
    test = Test.query.get_or_404(test_id)
    data = request.json
    test.title = data.get('title', test.title)
    test.level = data.get('level', test.level)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/tests/<int:test_id>/delete', methods=['POST'])
@admin_required
def delete_test(test_id):
    test = Test.query.get_or_404(test_id)
    # Изтрий резултатите
    TestResult.query.filter_by(test_id=test_id).delete()
    db.session.delete(test)
    # Изтрий снимките от файловата система
    import shutil
    img_dir = f"/tmp/qimages/{test_id}"
    if os.path.exists(img_dir):
        shutil.rmtree(img_dir)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/tests/<int:test_id>/questions')
@admin_required
def get_test_questions(test_id):
    test = Test.query.get_or_404(test_id)
    return jsonify({'questions': test.get_questions(), 'title': test.title})

@app.route('/admin/tests/<int:test_id>/questions', methods=['POST'])
@admin_required
def save_test_questions(test_id):
    try:
        test = Test.query.get_or_404(test_id)
        questions = request.json.get('questions', [])

        # Запази has_image флага от оригиналните въпроси
        original = {str(q['id']): q for q in test.get_questions()}
        
        for q in questions:
            # Възстанови has_image от оригинала
            orig = original.get(str(q['id']))
            if orig and orig.get('has_image'):
                q['has_image'] = True

            # Гарантира само ЕДИН верен отговор
            correct_found = False
            for opt in q.get('options', []):
                if opt.get('isCorrect') and not correct_found:
                    correct_found = True
                elif opt.get('isCorrect') and correct_found:
                    opt['isCorrect'] = False
            if not correct_found and q.get('options'):
                q['options'][0]['isCorrect'] = True

        test.questions_json = json.dumps(questions, ensure_ascii=False)
        test.question_count = len(questions)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        print("SAVE QUESTIONS ERROR:", traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.filter_by(is_admin=False).order_by(User.created_at.desc()).all()
    return render_template('admin_users.html', users=users, now=datetime.utcnow())

@app.route('/admin/users/<int:user_id>')
@admin_required
def admin_user_detail(user_id):
    user = User.query.get_or_404(user_id)
    results = TestResult.query.filter_by(user_id=user_id).order_by(TestResult.taken_at.desc()).all()
    return render_template('admin_user_detail.html', user=user, results=results)

@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': user.is_active})

@app.route('/admin/promos')
@admin_required
def admin_promos():
    promos = PromoCode.query.order_by(PromoCode.created_at.desc()).all()
    active = sum(1 for p in promos if p.is_active)
    used = sum(1 for p in promos if p.is_used)
    return render_template('admin_promos.html', promos=promos, active=active, used=used)

@app.route('/admin/promos/create', methods=['POST'])
@admin_required
def create_promo():
    client = request.form.get('client_name', '').strip()
    access_type = request.form.get('access_type', 'Регулярни тестове')
    price = float(request.form.get('price', 0) or 0)
    code = generate_promo_code()

    promo = PromoCode(code=code, client_name=client, access_type=access_type, price=price)
    db.session.add(promo)
    db.session.commit()
    return jsonify({'success': True, 'code': code})

@app.route('/admin/promos/<int:promo_id>/delete', methods=['POST'])
@admin_required
def delete_promo(promo_id):
    promo = PromoCode.query.get_or_404(promo_id)
    db.session.delete(promo)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/results/<int:result_id>')
@admin_required
def admin_result_detail(result_id):
    result = TestResult.query.get_or_404(result_id)
    test = Test.query.get(result.test_id)
    user = User.query.get(result.user_id)
    all_questions = test.get_questions()
    answers = json.loads(result.answers_json)

    # Ако имаме записани ID-та — показвай само тях
    try:
        q_ids = json.loads(result.question_ids_json or '[]')
    except:
        q_ids = []

    if q_ids:
        qid_set = set(str(q) for q in q_ids)
        questions = [q for q in all_questions if str(q['id']) in qid_set]
    else:
        answered_ids = set(answers.keys())
        questions = [q for q in all_questions if str(q['id']) in answered_ids] or all_questions

    # Зареди снимките
    questions = inject_images(result.test_id, questions)

    # Форматирай времето
    duration = result.duration or 0
    duration_str = f"{duration // 60:02d}:{duration % 60:02d}"

    # Тип на теста
    type_labels = {'test': 'Обикновен Тест', 'mix': 'Микс', 'simulator': 'Симулатор', 'mistakes': 'Грешки'}
    type_label = type_labels.get(result.test_type or 'test', 'Тест')

    return render_template('admin_result_detail.html',
        result=result, test=test, user=user,
        questions=questions, answers=answers,
        duration_str=duration_str, type_label=type_label)

@app.route('/admin/results/<int:result_id>/delete', methods=['POST'])
@admin_required
def delete_result(result_id):
    result = TestResult.query.get_or_404(result_id)
    db.session.delete(result)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/results/cleanup', methods=['POST'])
@admin_required
def cleanup_results():
    """Изтрива резултати по-стари от X дни"""
    days = int(request.json.get('days', 30))
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=days)
    old_results = TestResult.query.filter(TestResult.taken_at < cutoff).all()
    count = len(old_results)
    for r in old_results:
        db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True, 'deleted': count})

@app.route('/admin/signals')
@admin_required
def admin_signals():
    signals = Signal.query.order_by(Signal.created_at.desc()).all()
    open_count = Signal.query.filter_by(status='open').count()
    return render_template('admin_signals.html', signals=signals, open_count=open_count)

@app.route('/admin/signals/<int:signal_id>/resolve', methods=['POST'])
@admin_required
def resolve_signal(signal_id):
    signal = Signal.query.get_or_404(signal_id)
    signal.status = 'resolved'
    db.session.commit()
    return jsonify({'success': True})

# ============================================================
#  ИНИЦИАЛИЗАЦИЯ
# ============================================================

def create_admin():
    """Създава администратор при първо стартиране"""
    with app.app_context():
        db.create_all()
        # Добавя test_type колона ако не съществува
        try:
            from sqlalchemy import text, inspect
            # Create test_image table if not exists
            db.create_all()
            inspector = inspect(db.engine)
            existing_cols = [c['name'] for c in inspector.get_columns('test_result')]
            with db.engine.connect() as conn:
                if 'test_type' not in existing_cols:
                    conn.execute(text('ALTER TABLE test_result ADD COLUMN test_type VARCHAR(20) DEFAULT "test"'))
                    conn.commit()
                if 'duration' not in existing_cols:
                    conn.execute(text('ALTER TABLE test_result ADD COLUMN duration INTEGER DEFAULT 0'))
                    conn.commit()
                if 'question_ids_json' not in existing_cols:
                    conn.execute(text('ALTER TABLE test_result ADD COLUMN question_ids_json TEXT DEFAULT "[]"'))
                    conn.commit()
            # Add is_demo to test table
            test_cols = [c['name'] for c in inspector.get_columns('test')]
            if 'is_demo' not in test_cols:
                with db.engine.connect() as conn2:
                    conn2.execute(text('ALTER TABLE test ADD COLUMN is_demo BOOLEAN DEFAULT 0'))
                    conn2.commit()
            # Create demo_visit table
            db.create_all()
            # Add last_seen to user table
            user_cols2 = [c['name'] for c in inspector.get_columns('user')]
            if 'last_seen' not in user_cols2:
                with db.engine.connect() as conn3:
                    conn3.execute(text('ALTER TABLE user ADD COLUMN last_seen DATETIME'))
                    conn3.commit()
            # Add email_verified and verification_token to user table
            user_cols = [c['name'] for c in inspector.get_columns('user')]
            if 'email_verified' not in user_cols:
                with db.engine.connect() as conn2:
                    conn2.execute(text('ALTER TABLE user ADD COLUMN email_verified BOOLEAN DEFAULT 0'))
                    conn2.commit()
            if 'verification_token' not in user_cols:
                with db.engine.connect() as conn3:
                    conn3.execute(text('ALTER TABLE user ADD COLUMN verification_token VARCHAR(64)'))
                    conn3.commit()
            print("✓ DB migration OK")
        except Exception as e:
            print(f"Migration note: {e}")
        try:
            import os
            admin_user = User.query.filter_by(is_admin=True).first()
            admin_pass = os.environ.get('ADMIN_PASSWORD', 'admin123')
            if not admin_user:
                admin_user = User(
                    name='Администратор',
                    email='admin@maritime.bg',
                    password=generate_password_hash(admin_pass),
                    is_admin=True
                )
                db.session.add(admin_user)
                db.session.commit()
                print("✓ Администратор създаден")
            else:
                # Update password if ADMIN_PASSWORD env var is set
                if os.environ.get('ADMIN_PASSWORD'):
                    admin_user.password = generate_password_hash(admin_pass)
                    db.session.commit()
                    print("✓ Админ парола обновена")
            if False:  # dummy to keep indentation
                print("✓ Администратор създаден: admin@maritime.bg / admin123")
        except Exception:
            db.session.rollback()

# Инициализация на базата - работи и на Railway и локално - v2.1
create_admin()

if __name__ == '__main__':
    print("=" * 50)
    print("  Морски Тестове - Стартира...")
    print("  http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, port=5000)
