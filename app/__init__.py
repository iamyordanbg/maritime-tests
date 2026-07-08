import os
from flask import Flask
from .extensions import db
from .routes import register_blueprints
from .models import User, Test, TestImage, DemoVisit, TestResult, PromoCode, MonthlySnapshot, Signal, Post, PostComment, FreeSession
from config import config
from werkzeug.security import generate_password_hash

def create_app(config_name=None):
    if config_name is None:
        config_name = os.environ.get("FLASK_ENV", "production")

    app = Flask(__name__, 
                template_folder="templates",
                static_folder="static")
    app.config.from_object(config[config_name])

    # Кара браузъра да пази снимките дълго локално, вместо да пита сървъра
    # "промениха ли се?" при всяко зареждане на страница — за тест с много
    # снимки, това означава реални мрежови "разговора" при ВСЯКО повторно
    # отваряне на СЪЩИЯ тест, дори снимките изобщо да не са се променили.
    # 30 дни е разумен баланс — снимките към въпроси не се сменят често.
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 60 * 60 * 24 * 30

    # Инициализираме extensions
    db.init_app(app)

    # Регистрираме blueprints
    register_blueprints(app)

    # Context processors
    @app.context_processor
    def inject_greeting():
        from flask import session
        try:
            just_logged_in = session.pop("just_logged_in", False)
        except Exception:
            just_logged_in = False
        return dict(just_logged_in=just_logged_in)

    @app.context_processor
    def inject_admin_user():
        from flask import session
        admin_user = None
        try:
            if session.get("is_admin") and session.get("user_id"):
                admin_user = User.query.get(session["user_id"])
        except Exception:
            pass
        return dict(admin_user=admin_user)

    @app.context_processor
    def inject_recaptcha():
        return dict(recaptcha_site_key=app.config.get("RECAPTCHA_SITE_KEY", ""))

    @app.before_request
    def _enforce_inactivity_timeout():
        """
        Автоматичен logout след N минути реална неактивност. /ping НЕ се брои
        за активност (той е keep-alive от JS таймер, не истинско действие на
        потребителя) — иначе сесията никога не би могла да изтече сама.
        """
        from flask import session as _session, request as _request
        from datetime import datetime as _dt
        if _request.endpoint in ('auth.ping', 'dashboard.ping') or _request.path.startswith('/static'):
            return

        if 'user_id' in _session:
            now_ts = _dt.utcnow().timestamp()
            last = _session.get('last_activity')
            timeout_seconds = app.config.get('INACTIVITY_TIMEOUT_MINUTES', 30) * 60

            if last and (now_ts - last) > timeout_seconds:
                _session.clear()
                if _request.path.startswith('/api/') or _request.endpoint in ('dashboard.support_unread',):
                    from flask import jsonify as _jsonify
                    return _jsonify({'error': 'session_expired'}), 401
                from flask import redirect as _redirect, url_for as _url_for
                return _redirect(_url_for('auth.login'))

            _session['last_activity'] = now_ts

    @app.context_processor
    def inject_plans():
        # billing/plans.html разчита на 'plans' — прави се include от sidebar модала
        # на много страници, не само от /billing/plans route-а, затова трябва да е
        # глобално достъпна, а не подавана ръчно от всеки отделен route.
        from app.services.plans import PLANS
        return dict(plans=PLANS)

    @app.context_processor
    def inject_now_and_usage():
        import math
        from datetime import datetime
        from flask import session
        now = datetime.utcnow()
        usage_days_left = 0
        try:
            if session.get("user_id"):
                u = User.query.get(session["user_id"])
                if u and u.plan in ('basic', 'plus', 'gold') and u.plan_expires_at:
                    secs = (u.plan_expires_at - now).total_seconds()
                    usage_days_left = max(0, math.ceil(secs / 86400))
        except Exception:
            pass
        return dict(now=now, usage_days_left=usage_days_left)

    with app.app_context():
        _migrate_db(app)
        _create_admin(app)
        _create_test_user(app)

        # Изтриваме само orphan test_result редове с user_id = NULL
        try:
            from sqlalchemy import text as _text
            with db.engine.connect() as _conn:
                _conn.execute(_text('DELETE FROM test_result WHERE user_id IS NULL'))
                _conn.commit()
        except Exception as e:
            print(f"✗ Cleanup error: {e}", flush=True)

    return app


def parse_xls_colors(filepath):
    """Чете XLS/XLSX и открива верни отговори по цвят на шрифта"""
    OPT_LETTERS = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
    questions = []

    if filepath.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Извличаме снимките — всяка снимка принадлежи на въпроса НАД нея
        image_map = {}
        try:
            all_images = ws._images
            print(f"PARSE: Found {len(all_images)} images in worksheet")
        except Exception as e:
            print(f"PARSE: Cannot access _images: {e}")
            all_images = []
        
        for img in all_images:
            try:
                anchor_row_0idx = img.anchor._from.row
                ws_row_of_image = anchor_row_0idx + 1
                question_ws_row = ws_row_of_image - 1
                # Try different methods to get image data
                try:
                    img_data = img._data()
                except:
                    try:
                        img_data = img.ref.blob
                    except:
                        img_data = bytes(img.ref._data)
                fmt = 'jpg' if img_data[:2] == b'\xff\xd8' else 'png'
                image_map[question_ws_row] = (img_data, fmt)
            except Exception as e:
                print(f"PARSE: Image error: {e}")

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            q_cell = row[0]
            if not q_cell.value or str(q_cell.value).strip() == '':
                continue
            q_text = str(q_cell.value).strip()
            options = []
            opt_idx = 0

            for cell in row[1:]:
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                is_correct = False
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == 'rgb':
                        rgb = color.rgb
                        if rgb not in ('00000000', 'FF000000', '000000'):
                            is_correct = True
                    elif color.type == 'theme':
                        if color.theme not in (0, 1):
                            is_correct = True
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            q_id = len(questions) + 1  # 1, 2, 3... последователно
            q = {'id': q_id, 'question': q_text, 'options': options}
            if r_idx in image_map:
                q['has_image'] = True
                q['_image_data'] = image_map[r_idx]
            questions.append(q)

    else:
        # XLS - използваме xlrd
        BLACK_IDX = 8
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        for r in range(1, ws.nrows):
            q_val = ws.cell(r, 0).value
            if not q_val or str(q_val).strip() == '':
                continue
            q_text = str(q_val).strip()
            options = []
            opt_idx = 0

            for c in range(1, ws.ncols):
                cell = ws.cell(r, c)
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                xf_idx = ws.cell_xf_index(r, c)
                xf = wb.xf_list[xf_idx]
                font = wb.font_list[xf.font_index]
                is_correct = (font.colour_index != BLACK_IDX)
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            questions.append({'id': len(questions) + 1, 'question': q_text, 'options': options})

    return questions

def generate_promo_code(prefix='MAR'):
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"{prefix}-{suffix}"

# ============================================================
#  AUTH ROUTES
# ============================================================

def _migrate_db(app):
    """Автоматична DB миграция"""
    from sqlalchemy import inspect, text
    with app.app_context():
        # За Postgres с няколко gunicorn worker-а: всеки процес независимо вика
        # create_app() -> _migrate_db(), значи 2+ процеса могат едновременно да
        # опитат СЪЩИТЕ ALTER TABLE заявки — точно причината за замръзналия deploy
        # (взаимно чакане на lock). Advisory lock гарантира, че само ЕДИН worker
        # реално мигрира; останалите прескачат веднага (не чакат, не се бъркат).
        is_postgres = db.engine.dialect.name == 'postgresql'
        lock_conn = None
        got_lock = True
        if is_postgres:
            try:
                lock_conn = db.engine.connect()
                got_lock = bool(lock_conn.execute(text("SELECT pg_try_advisory_lock(727384910)")).scalar())
            except Exception:
                got_lock = True  # ако locking логиката гръмне, не блокираме стартирането

        if not got_lock:
            print("⏭  Друг worker вече мигрира базата — прескачам.")
            if lock_conn:
                lock_conn.close()
            return

        try:
            db.create_all()
            inspector = inspect(db.engine)

            # User колони
            user_cols = [c["name"] for c in inspector.get_columns("user")]
            migrations = [
                ("plan", 'ALTER TABLE "user" ADD COLUMN plan VARCHAR(20) DEFAULT \'free\''),
                ("plan_update", 'UPDATE "user" SET plan = \'free\' WHERE plan IS NULL'),
                ("nick", 'ALTER TABLE "user" ADD COLUMN nick VARCHAR(100) DEFAULT \'\''),
                ("fullname", 'ALTER TABLE "user" ADD COLUMN fullname VARCHAR(100) DEFAULT \'\''),
                ("notif_subscription", 'ALTER TABLE "user" ADD COLUMN notif_subscription BOOLEAN DEFAULT 1'),
                ("email_verified", 'ALTER TABLE "user" ADD COLUMN email_verified BOOLEAN DEFAULT 0'),
                ("email_verified_fix", 'UPDATE "user" SET email_verified = TRUE WHERE email_verified = FALSE AND id > 0'),
                ("google_id", 'ALTER TABLE "user" ADD COLUMN google_id VARCHAR(200)'),
                ("last_seen", 'ALTER TABLE "user" ADD COLUMN last_seen TIMESTAMP'),
                ("is_active", 'ALTER TABLE "user" ADD COLUMN is_active BOOLEAN DEFAULT 0'),
                ("is_admin", 'ALTER TABLE "user" ADD COLUMN is_admin BOOLEAN DEFAULT 0'),
                ("library_test_id", 'ALTER TABLE "user" ADD COLUMN library_test_id INTEGER'),
                ("library_selected_at", 'ALTER TABLE "user" ADD COLUMN library_selected_at TIMESTAMP'),
                ("library_last_simulator_at", 'ALTER TABLE "user" ADD COLUMN library_last_simulator_at TIMESTAMP'),
                ("lifetime_test_count", 'ALTER TABLE "user" ADD COLUMN lifetime_test_count INTEGER DEFAULT 0'),
                ("plan_activated_at", 'ALTER TABLE "user" ADD COLUMN plan_activated_at TIMESTAMP'),
                ("plan_expires_at", 'ALTER TABLE "user" ADD COLUMN plan_expires_at TIMESTAMP'),
                ("gold_test_ids", 'ALTER TABLE "user" ADD COLUMN gold_test_ids TEXT'),
                ("plan_grace_until", 'ALTER TABLE "user" ADD COLUMN plan_grace_until TIMESTAMP'),
            ]
            with db.engine.connect() as conn:
                # Създаваме таблица за приложени миграции
                try:
                    conn.execute(text('CREATE TABLE IF NOT EXISTS _applied_migrations (name VARCHAR(100) PRIMARY KEY)'))
                    conn.commit()
                except Exception:
                    pass
                applied = set()
                try:
                    rows = conn.execute(text('SELECT name FROM _applied_migrations')).fetchall()
                    applied = {r[0] for r in rows}
                except Exception:
                    pass

                for col, sql in migrations:
                    # ALTER TABLE — само ако колоната липсва
                    if sql.strip().upper().startswith('ALTER') and col in user_cols:
                        continue
                    # UPDATE/INSERT — само ако не е вече приложена
                    if not sql.strip().upper().startswith('ALTER') and col in applied:
                        continue
                    try:
                        conn.execute(text(sql))
                        conn.commit()
                        if not sql.strip().upper().startswith('ALTER'):
                            conn.execute(text('INSERT INTO _applied_migrations (name) VALUES (:n) ON CONFLICT DO NOTHING'), {'n': col})
                            conn.commit()
                    except Exception:
                            pass

            # MonthlySnapshot таблица
            if "monthly_snapshot" not in inspector.get_table_names():
                MonthlySnapshot.__table__.create(db.engine)

            # Signal колони
            if 'signal' in inspector.get_table_names():
                sig_cols = [c['name'] for c in inspector.get_columns('signal')]
                with db.engine.connect() as conn:
                    for col, sql in [
                        ('user_email', 'ALTER TABLE signal ADD COLUMN user_email VARCHAR(120) DEFAULT \'\''),
                        ('reply', 'ALTER TABLE signal ADD COLUMN reply VARCHAR(500)'),
                        ('replied_at', 'ALTER TABLE signal ADD COLUMN replied_at TIMESTAMP'),
                        ('is_read', 'ALTER TABLE signal ADD COLUMN is_read BOOLEAN DEFAULT 0'),
                    ]:
                        if col not in sig_cols:
                            try:
                                conn.execute(text(sql))
                                conn.commit()
                            except Exception:
                                pass

            # TestImage колони — R2 поддръжка. image_data става nullable
            # (снимките в R2 не пазят base64 в базата, само reference),
            # storage маркира откъде да се чете ('db' старите, 'r2' новите),
            # r2_key пази пътя в bucket-a за trigger при delete.
            if 'test_image' in inspector.get_table_names():
                ti_cols = [c['name'] for c in inspector.get_columns('test_image')]
                with db.engine.connect() as conn:
                    for col, sql in [
                        ('storage', "ALTER TABLE test_image ADD COLUMN storage VARCHAR(10) DEFAULT 'db'"),
                        ('r2_key', 'ALTER TABLE test_image ADD COLUMN r2_key VARCHAR(255)'),
                    ]:
                        if col not in ti_cols:
                            try:
                                conn.execute(text(sql))
                                conn.commit()
                            except Exception:
                                pass
                    try:
                        conn.execute(text('ALTER TABLE test_image ALTER COLUMN image_data DROP NOT NULL'))
                        conn.commit()
                    except Exception:
                        pass

            # Ticket таблици
            if 'ticket' not in inspector.get_table_names():
                from app.models.ticket import Ticket, TicketMessage
                Ticket.__table__.create(db.engine)
                TicketMessage.__table__.create(db.engine)

            # tests_used колона в user
            user_cols2 = [c['name'] for c in inspector.get_columns('user')]
            if 'tests_used' not in user_cols2:
                with db.engine.connect() as conn:
                    try:
                        conn.execute(text('ALTER TABLE "user" ADD COLUMN tests_used INTEGER DEFAULT 0'))
                        conn.commit()
                    except Exception:
                        pass

            # pref_q_font_size / pref_a_font_size / pref_theme /
            # pref_q_font_family / pref_a_font_family - настройки за четене
            # на тестовете (менюто с 3-те чертички в хедъра), отделни
            # слайдери (0-10) и шрифтове за въпрос vs отговори
            user_cols_prefs = [c['name'] for c in inspector.get_columns('user')]
            with db.engine.connect() as conn:
                for col, sql in [
                    ('pref_q_font_size', "ALTER TABLE \"user\" ADD COLUMN pref_q_font_size INTEGER DEFAULT 5"),
                    ('pref_a_font_size', "ALTER TABLE \"user\" ADD COLUMN pref_a_font_size INTEGER DEFAULT 5"),
                    ('pref_highlight_intensity', "ALTER TABLE \"user\" ADD COLUMN pref_highlight_intensity INTEGER DEFAULT 5"),
                    ('pref_theme', "ALTER TABLE \"user\" ADD COLUMN pref_theme VARCHAR(10) DEFAULT 'dark'"),
                    ('pref_q_font_family', "ALTER TABLE \"user\" ADD COLUMN pref_q_font_family VARCHAR(20) DEFAULT 'default'"),
                    ('pref_a_font_family', "ALTER TABLE \"user\" ADD COLUMN pref_a_font_family VARCHAR(20) DEFAULT 'default'"),
                    ('pref_q_bold', "ALTER TABLE \"user\" ADD COLUMN pref_q_bold BOOLEAN DEFAULT TRUE"),
                    ('pref_a_bold', "ALTER TABLE \"user\" ADD COLUMN pref_a_bold BOOLEAN DEFAULT FALSE"),
                ]:
                    if col not in user_cols_prefs:
                        try:
                            conn.execute(text(sql))
                            conn.commit()
                        except Exception:
                            pass

            # tests_used колона в user
            user_cols_ext = [c['name'] for c in inspector.get_columns('user')]
            if 'tests_used' not in user_cols_ext:
                with db.engine.connect() as conn:
                    try:
                        conn.execute(text('ALTER TABLE "user" ADD COLUMN tests_used INTEGER DEFAULT 0'))
                        conn.commit()
                    except Exception:
                        pass

            # Payment колони
            if 'payment' in inspector.get_table_names():
                pay_cols = [c['name'] for c in inspector.get_columns('payment')]
                with db.engine.connect() as conn:
                    for col, sql in [
                        ('stripe_fee', 'ALTER TABLE payment ADD COLUMN stripe_fee FLOAT'),
                        ('net_amount', 'ALTER TABLE payment ADD COLUMN net_amount FLOAT'),
                    ]:
                        if col not in pay_cols:
                            try:
                                conn.execute(text(sql))
                                conn.commit()
                            except Exception:
                                pass

            # PromoCode колони (Gold активация)
            if 'promo_code' in inspector.get_table_names():
                promo_cols = [c['name'] for c in inspector.get_columns('promo_code')]
                with db.engine.connect() as conn:
                    for col, sql in [
                        ('plan', "ALTER TABLE promo_code ADD COLUMN plan VARCHAR(20) DEFAULT 'gold'"),
                        ('department', 'ALTER TABLE promo_code ADD COLUMN department VARCHAR(10)'),
                        ('level', 'ALTER TABLE promo_code ADD COLUMN level VARCHAR(50)'),
                        ('selected_test_ids', 'ALTER TABLE promo_code ADD COLUMN selected_test_ids TEXT'),
                        ('mistakes_grace_days', 'ALTER TABLE promo_code ADD COLUMN mistakes_grace_days INTEGER DEFAULT 60'),
                        ('activated_at', 'ALTER TABLE promo_code ADD COLUMN activated_at TIMESTAMP'),
                        ('shared_to', 'ALTER TABLE promo_code ADD COLUMN shared_to VARCHAR(120)'),
                        ('shared_at', 'ALTER TABLE promo_code ADD COLUMN shared_at TIMESTAMP'),
                        ('shared_count', 'ALTER TABLE promo_code ADD COLUMN shared_count INTEGER DEFAULT 0'),
                    ]:
                        if col not in promo_cols:
                            try:
                                conn.execute(text(sql))
                                conn.commit()
                            except Exception:
                                pass

            # Индекси за често филтрираните колони — без тях, заявки към растящи
            # таблици (TestResult расте с всеки решен тест) правят пълно
            # последователно сканиране вместо бърз индексиран lookup. CREATE
            # INDEX IF NOT EXISTS е безопасно — не пипа данни, само ускорява четенето.
            index_statements = [
                'CREATE INDEX IF NOT EXISTS ix_test_result_user_id ON test_result (user_id)',
                'CREATE INDEX IF NOT EXISTS ix_test_result_test_id ON test_result (test_id)',
                'CREATE INDEX IF NOT EXISTS ix_test_result_taken_at ON test_result (taken_at)',
                'CREATE INDEX IF NOT EXISTS ix_gold_grant_user_id ON gold_grant (user_id)',
                'CREATE INDEX IF NOT EXISTS ix_plan_grant_user_id ON plan_grant (user_id)',
                'CREATE INDEX IF NOT EXISTS ix_ticket_user_id ON ticket (user_id)',
                'CREATE INDEX IF NOT EXISTS ix_ticket_message_ticket_id ON ticket_message (ticket_id)',
            ]
            with db.engine.connect() as conn:
                for sql in index_statements:
                    try:
                        conn.execute(text(sql))
                        conn.commit()
                    except Exception:
                        pass

            print("✓ DB migration OK")
        except Exception as e:
            print(f"Migration error: {e}")
        finally:
            if is_postgres and lock_conn:
                try:
                    lock_conn.execute(text("SELECT pg_advisory_unlock(727384910)"))
                except Exception:
                    pass
                lock_conn.close()



def _create_admin(app):
    """Създава администратор и прави DB миграция"""
    with app.app_context():
        from sqlalchemy import text, inspect
        from werkzeug.security import generate_password_hash
        import os

        db.create_all()
        try:
            inspector = inspect(db.engine)

            # test_result колони
            existing_cols = [c['name'] for c in inspector.get_columns('test_result')]
            with db.engine.connect() as conn:
                for col, sql in [
                    ('test_type', 'ALTER TABLE test_result ADD COLUMN test_type VARCHAR(20) DEFAULT \'test\''),
                    ('duration', 'ALTER TABLE test_result ADD COLUMN duration INTEGER DEFAULT 0'),
                    ('question_ids_json', 'ALTER TABLE test_result ADD COLUMN question_ids_json TEXT DEFAULT \'[]\''),
                    ('user_seq', 'ALTER TABLE test_result ADD COLUMN user_seq INTEGER'),
                ]:
                    if col not in existing_cols:
                        try:
                            conn.execute(text(sql))
                            conn.commit()
                        except Exception:
                            pass

            # Изтриваме orphan test_result редове с user_id = null
            with db.engine.connect() as conn:
                try:
                    conn.execute(text('DELETE FROM test_result WHERE user_id IS NULL'))
                    conn.commit()
                except Exception:
                    pass

            # Еднократен backfill: попълва user_seq за СЪЩЕСТВУВАЩИ резултати
            # (нови NULL колони при първо добавяне) - пореден номер ПО
            # ПОТРЕБИТЕЛ, хронологично по taken_at. Postgres има вграден
            # ROW_NUMBER() window function - работи директно. При грешка
            # (напр. локален SQLite при разработка) тихо се прескача.
            with db.engine.connect() as conn:
                try:
                    conn.execute(text('CREATE TABLE IF NOT EXISTS _applied_migrations (name VARCHAR(100) PRIMARY KEY)'))
                    conn.commit()
                except Exception:
                    pass
                already_backfilled = False
                try:
                    row = conn.execute(text("SELECT 1 FROM _applied_migrations WHERE name = 'user_seq_backfill'")).fetchone()
                    already_backfilled = row is not None
                except Exception:
                    pass
                if not already_backfilled:
                    try:
                        conn.execute(text('''
                            UPDATE test_result
                            SET user_seq = sub.rn
                            FROM (
                                SELECT id, ROW_NUMBER() OVER (PARTITION BY user_id ORDER BY taken_at) AS rn
                                FROM test_result
                            ) AS sub
                            WHERE test_result.id = sub.id AND test_result.user_seq IS NULL
                        '''))
                        # user.lifetime_test_count = максималния user_seq, който вече има
                        conn.execute(text('''
                            UPDATE "user"
                            SET lifetime_test_count = COALESCE((
                                SELECT MAX(user_seq) FROM test_result WHERE test_result.user_id = "user".id
                            ), 0)
                        '''))
                        conn.execute(text("INSERT INTO _applied_migrations (name) VALUES ('user_seq_backfill') ON CONFLICT DO NOTHING"))
                        conn.commit()
                        print("✓ user_seq backfill приложен")
                    except Exception as _e:
                        conn.rollback()
                        print(f"⚠ user_seq backfill пропуснат ({_e})")

            # test колони
            test_cols = [c['name'] for c in inspector.get_columns('test')]
            if 'is_demo' not in test_cols:
                with db.engine.connect() as conn:
                    try:
                        conn.execute(text('ALTER TABLE test ADD COLUMN is_demo BOOLEAN DEFAULT 0'))
                        conn.commit()
                    except Exception:
                        pass

            # test_image колони — снимките вече се пазят в базата (persistent),
            # не на диска на контейнера (ephemeral, изтрива се при redeploy)
            if 'test_image' in inspector.get_table_names():
                ti_cols = [c['name'] for c in inspector.get_columns('test_image')]
                if 'format' not in ti_cols:
                    with db.engine.connect() as conn:
                        try:
                            conn.execute(text("ALTER TABLE test_image ADD COLUMN format VARCHAR(10) DEFAULT 'jpg'"))
                            conn.commit()
                        except Exception:
                            pass
                with db.engine.connect() as conn:
                    try:
                        conn.execute(text(
                            'CREATE UNIQUE INDEX IF NOT EXISTS ix_test_image_test_question '
                            'ON test_image (test_id, question_id)'
                        ))
                        conn.commit()
                    except Exception:
                        pass

            # user колони
            user_cols = [c['name'] for c in inspector.get_columns('user')]
            with db.engine.connect() as conn:
                for col, sql in [
                    ('last_seen', 'ALTER TABLE "user" ADD COLUMN last_seen TIMESTAMP'),
                    ('email_verified', 'ALTER TABLE "user" ADD COLUMN email_verified BOOLEAN DEFAULT 0'),
                    ('verification_token', 'ALTER TABLE "user" ADD COLUMN verification_token VARCHAR(64)'),
                    ('otp_code', 'ALTER TABLE "user" ADD COLUMN otp_code VARCHAR(6)'),
                    ('otp_expires', 'ALTER TABLE "user" ADD COLUMN otp_expires TIMESTAMP'),
                    ('reset_token_expires', 'ALTER TABLE "user" ADD COLUMN reset_token_expires TIMESTAMP'),
                    ('notif_subscription', 'ALTER TABLE "user" ADD COLUMN notif_subscription BOOLEAN DEFAULT 1'),
                    ('firstname', 'ALTER TABLE "user" ADD COLUMN firstname VARCHAR(100) DEFAULT \'\''),
                    ('lastname', 'ALTER TABLE "user" ADD COLUMN lastname VARCHAR(100) DEFAULT \'\''),
                    ('nick', 'ALTER TABLE "user" ADD COLUMN nick VARCHAR(100) DEFAULT \'\''),
                    ('fullname', 'ALTER TABLE "user" ADD COLUMN fullname VARCHAR(100) DEFAULT \'\''),
                    ('google_id', 'ALTER TABLE "user" ADD COLUMN google_id VARCHAR(200)'),
                    ('is_active', 'ALTER TABLE "user" ADD COLUMN is_active BOOLEAN DEFAULT 0'),
                    ('library_test_id', 'ALTER TABLE "user" ADD COLUMN library_test_id INTEGER'),
                    ('library_selected_at', 'ALTER TABLE "user" ADD COLUMN library_selected_at TIMESTAMP'),
                    ('library_last_simulator_at', 'ALTER TABLE "user" ADD COLUMN library_last_simulator_at TIMESTAMP'),
                    ('lifetime_test_count', 'ALTER TABLE "user" ADD COLUMN lifetime_test_count INTEGER DEFAULT 0'),
                ]:
                    if col not in user_cols:
                        try:
                            conn.execute(text(sql))
                            conn.commit()
                        except Exception:
                            pass

            # MonthlySnapshot
            if 'monthly_snapshot' not in inspector.get_table_names():
                from app.models.snapshot import MonthlySnapshot
                MonthlySnapshot.__table__.create(db.engine)

            # Signal колони
            if 'signal' in inspector.get_table_names():
                sig_cols = [c['name'] for c in inspector.get_columns('signal')]
                with db.engine.connect() as conn:
                    for col, sql in [
                        ('user_email', 'ALTER TABLE signal ADD COLUMN user_email VARCHAR(120) DEFAULT \'\''),
                        ('reply', 'ALTER TABLE signal ADD COLUMN reply VARCHAR(500)'),
                        ('replied_at', 'ALTER TABLE signal ADD COLUMN replied_at TIMESTAMP'),
                        ('is_read', 'ALTER TABLE signal ADD COLUMN is_read BOOLEAN DEFAULT 0'),
                    ]:
                        if col not in sig_cols:
                            try:
                                conn.execute(text(sql))
                                conn.commit()
                            except Exception:
                                pass

            # Ticket таблици
            if 'ticket' not in inspector.get_table_names():
                from app.models.ticket import Ticket, TicketMessage
                Ticket.__table__.create(db.engine)
                TicketMessage.__table__.create(db.engine)

            # tests_used колона в user
            user_cols2 = [c['name'] for c in inspector.get_columns('user')]
            if 'tests_used' not in user_cols2:
                with db.engine.connect() as conn:
                    try:
                        conn.execute(text('ALTER TABLE "user" ADD COLUMN tests_used INTEGER DEFAULT 0'))
                        conn.commit()
                    except Exception:
                        pass

            # Payment колони
            if 'payment' in inspector.get_table_names():
                pay_cols = [c['name'] for c in inspector.get_columns('payment')]
                with db.engine.connect() as conn:
                    for col, sql in [
                        ('stripe_fee', 'ALTER TABLE payment ADD COLUMN stripe_fee FLOAT'),
                        ('net_amount', 'ALTER TABLE payment ADD COLUMN net_amount FLOAT'),
                        ('promo_email_sent', 'ALTER TABLE payment ADD COLUMN promo_email_sent BOOLEAN DEFAULT FALSE'),
                        ('promo_email_sent_at', 'ALTER TABLE payment ADD COLUMN promo_email_sent_at TIMESTAMP'),
                    ]:
                        if col not in pay_cols:
                            try:
                                conn.execute(text(sql))
                                conn.commit()
                            except Exception:
                                pass

            print("✓ DB migration OK")
        except Exception as e:
            print(f"Migration note: {e}")

        # Създаваме admin акаунт
        try:
            from app.models.user import User
            admin_pass = os.environ.get('ADMIN_PASSWORD', 'admin123')
            admin_user = User.query.filter_by(is_admin=True).first()
            if not admin_user:
                admin_user = User(
                    name='Администратор',
                    email='admin@maritime.bg',
                    password=generate_password_hash(admin_pass),
                    is_admin=True,
                    email_verified=True
                )
                db.session.add(admin_user)
                db.session.commit()
                print("✓ Администратор създаден")
            elif os.environ.get('ADMIN_PASSWORD'):
                admin_user.password = generate_password_hash(admin_pass)
                db.session.commit()
        except Exception:
            db.session.rollback()


def _create_test_user(app):
    """Тестов потребител - само за разработка"""
    with app.app_context():
        from werkzeug.security import generate_password_hash
        from app.models.user import User
        try:
            test = User.query.filter_by(email='test@maritime.bg').first()
            if not test:
                test = User(
                    name='Test User',
                    email='test@maritime.bg',
                    password=generate_password_hash('test123'),
                    is_admin=False,
                    is_active=False,
                    email_verified=True,
                )
                db.session.add(test)
                db.session.commit()
                print("✓ Тестов потребител създаден: test@maritime.bg / test123")
            else:
                if test.is_active:
                    test.is_active = False
                    db.session.commit()
                    print("✓ Тестов потребител върнат на Free план")
        except Exception as e:
            db.session.rollback()
            print(f"Test user error: {e}")
