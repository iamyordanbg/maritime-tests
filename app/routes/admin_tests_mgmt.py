"""
app/routes/admin_tests_mgmt.py
================================
Admin: Tests upload/edit/delete/questions — extraction-нат от admin.py
(Group A audit, File Limits).
"""
from flask import Blueprint, render_template, request, session, jsonify
from werkzeug.utils import secure_filename
from app.extensions import db
from app.models.test import Test
from app.models.result import TestResult
from app.utils.decorators import admin_required
from app.utils.images import inject_images, save_test_images, delete_test_images
import os, json
import tempfile
import xlrd

admin_tests_mgmt = Blueprint("admin_tests_mgmt", __name__, url_prefix="/admin")


def parse_xls_colors(filepath):
    """Чете XLS/XLSX и открива верни отговори по цвят на шрифта"""
    OPT_LETTERS = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
    questions = []

    if filepath.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Извличаме снимките — всяка снимка принадлежи на въпроса НАД нея
        image_map = {}
        try:
            all_images = ws._images
            print(f"PARSE: Found {len(all_images)} images in worksheet")
        except Exception as e:
            print(f"PARSE: Cannot access _images: {e}")
            all_images = []
        
        for img in all_images:
            try:
                # Различни начини снимката да е "закачена" в Excel файла
                # (зависи от версията/инструмента, с който е създаден файлът):
                # OneCellAnchor/TwoCellAnchor -> anchor._from.row (най-честият случай)
                # AbsoluteAnchor -> няма _from, използва pos (x,y в EMU) - трябва
                # да изчислим приблизителния ред от Y координатата.
                question_ws_row = None
                anchor = img.anchor
                if hasattr(anchor, '_from') and anchor._from is not None:
                    question_ws_row = anchor._from.row + 1
                elif hasattr(anchor, 'pos') and anchor.pos is not None:
                    # AbsoluteAnchor: pos.y е в EMU (914400 EMU = 1 инч).
                    # Стандартен row height ~20px ~15pt ~190500 EMU - грубо
                    # изчисление, по-добре от пълен пропуск на снимката.
                    EMU_PER_ROW_APPROX = 190500
                    question_ws_row = int(anchor.pos.y / EMU_PER_ROW_APPROX) + 1
                else:
                    print(f"PARSE: Image with unsupported anchor type: {type(anchor).__name__}, skipping")
                    continue

                # Try different methods to get image data
                try:
                    img_data = img._data()
                except:
                    try:
                        img_data = img.ref.blob
                    except:
                        img_data = bytes(img.ref._data)
                fmt = 'jpg' if img_data[:2] == b'\xff\xd8' else 'png'
                image_map[question_ws_row] = (img_data, fmt)
                print(f"PARSE: Image matched to worksheet row {question_ws_row} (anchor type: {type(anchor).__name__})")
            except Exception as e:
                print(f"PARSE: Image error (anchor type {type(getattr(img, 'anchor', None)).__name__}): {e}")

        if all_images and not image_map:
            print(f"PARSE: WARNING - {len(all_images)} images found in file but NONE could be matched to a question row. Check anchor type compatibility above.")

        q_rows_found = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            q_cell = row[0]
            if not q_cell.value or str(q_cell.value).strip() == '':
                continue
            q_rows_found.append(r_idx)
            q_text = str(q_cell.value).strip()
            options = []
            opt_idx = 0

            for cell in row[1:]:
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                is_correct = False
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == 'rgb':
                        rgb = color.rgb
                        if rgb not in ('00000000', 'FF000000', '000000'):
                            is_correct = True
                    elif color.type == 'theme':
                        if color.theme not in (0, 1):
                            is_correct = True
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            q_id = len(questions) + 1  # 1, 2, 3... последователно
            q = {'id': q_id, 'question': q_text, 'options': options}
            questions.append(q)

        print(f"PARSE: Question rows found: {q_rows_found}")
        print(f"PARSE: Image rows computed: {list(image_map.keys())}")

        # Свързваме снимките с въпросите по НАЙ-БЛИЗКИЯ ПРЕДХОЖДАЩ ред, не по
        # точно съвпадение - някои Excel файлове анкерират снимката на
        # СЪЩИЯ ред като въпроса (точно съвпадение работи), но други я
        # поставят на СЛЕДВАЩИЯ (празен откъм текст) ред под въпроса
        # (засечено реално: снимки на ред 3,5,7 между въпроси на ред
        # 2,4,6 - точното съвпадение никога не намираше нищо).
        import bisect
        for img_row, img_payload in image_map.items():
            idx = bisect.bisect_right(q_rows_found, img_row) - 1
            if idx >= 0:
                questions[idx]['has_image'] = True
                questions[idx]['_image_data'] = img_payload

    else:
        # XLS - използваме xlrd
        BLACK_IDX = 8
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        for r in range(1, ws.nrows):
            q_val = ws.cell(r, 0).value
            if not q_val or str(q_val).strip() == '':
                continue
            q_text = str(q_val).strip()
            options = []
            opt_idx = 0

            for c in range(1, ws.ncols):
                cell = ws.cell(r, c)
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                xf_idx = ws.cell_xf_index(r, c)
                xf = wb.xf_list[xf_idx]
                font = wb.font_list[xf.font_index]
                is_correct = (font.colour_index != BLACK_IDX)
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            questions.append({'id': len(questions) + 1, 'question': q_text, 'options': options})

    return questions

from app.utils.images import inject_images, save_test_images, delete_test_images


@admin_tests_mgmt.route('/tests/force-upload', methods=['POST'])
@admin_required
def force_upload():
    """Качва тест използвайки вече парснатите данни от сесията"""
    pending_file = session.get('pending_upload_file')
    if pending_file and __import__('os').path.exists(pending_file):
        with open(pending_file) as _pf:
            pending = __import__('json').load(_pf)
    else:
        pending = session.get('pending_upload')
    if not pending:
        return jsonify({'error': 'Няма данни за качване'}), 400
    
    new_title = request.json.get('title', pending['title'])
    
    test = Test(
        title=new_title,
        category=pending['category'],
        level=pending['level'],
        questions_json=pending['questions_json'],
        question_count=pending['question_count'],
        is_demo=False
    )
    db.session.add(test)
    db.session.flush()
    
    # Запази снимките — trайно в базата (или R2), не на диска на контейнера
    if pending.get('images_b64'):
        import base64
        images_to_save = [
            (qid, (base64.b64decode(b64data), fmt))
            for qid, b64data, fmt in pending['images_b64']
        ]
        print(f"FORCE_UPLOAD: Decoding {len(images_to_save)} pending images for test {test.id}")
        save_test_images(test.id, images_to_save)
    else:
        print("FORCE_UPLOAD: Няма pending снимки за този upload (нормално, ако тестът няма снимки)")

    db.session.commit()
    session.pop('pending_upload', None)
    return jsonify({'success': True, 'title': new_title, 'total': pending['question_count']})

# toggle_demo route removed - use /admin/demo/toggle/<id>

@admin_tests_mgmt.route('/tests/upload', methods=['POST'])
@admin_required
def upload_test():
    file = request.files.get('file')
    title = request.form.get('title', '').strip()
    category = request.form.get('category', 'deck')
    level = request.form.get('level', 'Operational Level')

    if not file:
        return jsonify({'error': 'Няма файл'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(tempfile.gettempdir(), filename)
    file.save(filepath)

    try:
        print(f"UPLOAD: Starting parse of {filename}, size={os.path.getsize(filepath)} bytes")
        questions = parse_xls_colors(filepath)
        print(f"UPLOAD: Parsed {len(questions)} questions")
        with_img = sum(1 for q in questions if q.get('has_image'))
        print(f"UPLOAD: Questions with images: {with_img}")
        final_title = title if title else filename.replace('.xls', '').replace('.xlsx', '')
        
        # Провери за дублиращо се заглавие
        existing = Test.query.filter_by(title=final_title).first()
        if existing:
            force = request.form.get('force', 'false')
            if force != 'true':
                # Запази парснатите данни в сесията за по-късно
                import pickle, base64
                # Запазваме в /tmp вместо в сесията (cookie limit)
                _pending_data = {
                    'questions_json': __import__('json').dumps(
                        [{k: v for k, v in q.items() if k != '_image_data'} for q in questions],
                        ensure_ascii=False
                    ),
                    'question_count': len(questions),
                    'category': category,
                    'level': level,
                    'title': final_title,
                    'images': [(q['id'], q['_image_data']) for q in questions if '_image_data' in q]
                }
                _pending_file = f'/tmp/pending_upload_{session.get("user_id","admin")}.json'
                with open(_pending_file, 'w') as _pf:
                    # ВАЖНО: images СА включени тук (base64), не изключени —
                    # преди тази поправка се губеха мълчаливо при force upload
                    # на тест със същото заглавие (никога не стигаха до
                    # save_test_images(), нямаше никаква грешка в логовете).
                    _pending_images_b64 = [
                        [qid, base64.b64encode(img_bytes).decode('ascii'), fmt]
                        for qid, (img_bytes, fmt) in _pending_data['images']
                    ]
                    __import__('json').dump({
                        **{k: v for k, v in _pending_data.items() if k != 'images'},
                        'images_b64': _pending_images_b64,
                    }, _pf)
                session['pending_upload_file'] = _pending_file
                session['pending_upload'] = {
                    'title': final_title,
                    'category': category,
                    'level': level,
                    'question_count': len(questions)
                }
                os.remove(filepath)
                return jsonify({'duplicate': True, 'title': final_title})
            else:
                # Намери следващия свободен индекс
                idx = 1
                while Test.query.filter_by(title=f"{final_title} ({idx})").first():
                    idx += 1
                final_title = f"{final_title} ({idx})" 

        # Извади снимките преди да запишем JSON
        images_to_save = []
        for q in questions:
            if '_image_data' in q:
                images_to_save.append((q['id'], q.pop('_image_data')))

        test = Test(
            title=final_title,
            category=category,
            level=level,
            questions_json=json.dumps(questions, ensure_ascii=False),
            question_count=len(questions),
            is_demo=False
        )
        db.session.add(test)
        db.session.flush()
        test_id_for_images = test.id
        db.session.commit()
        os.remove(filepath)

        # Запази снимките — trайно в базата, не на диска на контейнера
        if images_to_save:
            save_test_images(test_id_for_images, images_to_save)

        return jsonify({'success': True, 'total': len(questions), 'title': final_title})
    except Exception as e:
        try: os.remove(filepath)
        except: pass
        return jsonify({'error': str(e)}), 500

@admin_tests_mgmt.route('/tests/<int:test_id>/edit')
@admin_required
def edit_test(test_id):
    test = Test.query.get_or_404(test_id)
    questions = test.get_questions()
    questions = inject_images(test_id, questions)
    return render_template('admin/edit_test.html', test=test, questions=questions)

@admin_tests_mgmt.route('/tests/<int:test_id>/update-info', methods=['POST'])
@admin_required
def update_test_info(test_id):
    test = Test.query.get_or_404(test_id)
    data = request.json
    test.title = data.get('title', test.title)
    test.level = data.get('level', test.level)
    db.session.commit()
    return jsonify({'success': True})

@admin_tests_mgmt.route('/tests/<int:test_id>/delete', methods=['POST'])
@admin_required
def delete_test(test_id):
    test = Test.query.get_or_404(test_id)
    # ВАЖЕН РЕД: изтриваме децата (TestResult, TestImage - имат FK към
    # test.id) ПРЕДИ да маркираме самия Test за триене. SQLite (dev/test
    # среда) не налага FK ограничения по подразбиране и маскира бъга, но
    # production Postgres го прави строго - грешен ред води до FK
    # constraint violation -> 500 грешка (реален случай, засечен от
    # потребител при триене на тест със снимки).
    TestResult.query.filter_by(test_id=test_id).delete()
    delete_test_images(test_id)
    db.session.delete(test)
    db.session.commit()
    return jsonify({'success': True})

@admin_tests_mgmt.route('/tests/<int:test_id>/questions')
@admin_required
def get_test_questions(test_id):
    test = Test.query.get_or_404(test_id)
    return jsonify({'questions': test.get_questions(), 'title': test.title})

@admin_tests_mgmt.route('/tests/<int:test_id>/questions', methods=['POST'])
@admin_required
def save_test_questions(test_id):
    try:
        test = Test.query.get_or_404(test_id)
        questions = request.json.get('questions', [])

        # Запази has_image флага от оригиналните въпроси
        original = {str(q['id']): q for q in test.get_questions()}
        
        for q in questions:
            # Възстанови has_image от оригинала
            orig = original.get(str(q['id']))
            if orig and orig.get('has_image'):
                q['has_image'] = True

            # Гарантира само ЕДИН верен отговор
            correct_found = False
            for opt in q.get('options', []):
                if opt.get('isCorrect') and not correct_found:
                    correct_found = True
                elif opt.get('isCorrect') and correct_found:
                    opt['isCorrect'] = False
            if not correct_found and q.get('options'):
                q['options'][0]['isCorrect'] = True

        test.questions_json = json.dumps(questions, ensure_ascii=False)
        test.question_count = len(questions)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        print("SAVE QUESTIONS ERROR:", traceback.format_exc())
        return jsonify({'error': str(e)}), 500
