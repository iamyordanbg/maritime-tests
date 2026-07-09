from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify
from werkzeug.utils import secure_filename
from app.extensions import db
from app.models.user import User
from app.models.test import Test, TestImage, DemoVisit
from app.models.result import TestResult
from app.models.promo import PromoCode
from app.models.payment import Payment
from app.models.signal import Signal
from app.models.ticket import Ticket, TicketMessage
from app.models.snapshot import MonthlySnapshot
from app.services.stats import get_admin_stats, record_monthly_snapshot
from app.utils.decorators import admin_required
from datetime import datetime, timedelta
import os, json

admin = Blueprint("admin", __name__, url_prefix="/admin")

import tempfile
import xlrd

def parse_xls_colors(filepath):
    """Чете XLS/XLSX и открива верни отговори по цвят на шрифта"""
    OPT_LETTERS = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
    questions = []

    if filepath.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Извличаме снимките — всяка снимка принадлежи на въпроса НАД нея
        image_map = {}
        try:
            all_images = ws._images
            print(f"PARSE: Found {len(all_images)} images in worksheet")
        except Exception as e:
            print(f"PARSE: Cannot access _images: {e}")
            all_images = []
        
        for img in all_images:
            try:
                anchor_row_0idx = img.anchor._from.row
                ws_row_of_image = anchor_row_0idx + 1
                question_ws_row = ws_row_of_image - 1
                # Try different methods to get image data
                try:
                    img_data = img._data()
                except:
                    try:
                        img_data = img.ref.blob
                    except:
                        img_data = bytes(img.ref._data)
                fmt = 'jpg' if img_data[:2] == b'\xff\xd8' else 'png'
                image_map[question_ws_row] = (img_data, fmt)
            except Exception as e:
                print(f"PARSE: Image error: {e}")

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            q_cell = row[0]
            if not q_cell.value or str(q_cell.value).strip() == '':
                continue
            q_text = str(q_cell.value).strip()
            options = []
            opt_idx = 0

            for cell in row[1:]:
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                is_correct = False
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == 'rgb':
                        rgb = color.rgb
                        if rgb not in ('00000000', 'FF000000', '000000'):
                            is_correct = True
                    elif color.type == 'theme':
                        if color.theme not in (0, 1):
                            is_correct = True
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            q_id = len(questions) + 1  # 1, 2, 3... последователно
            q = {'id': q_id, 'question': q_text, 'options': options}
            if r_idx in image_map:
                q['has_image'] = True
                q['_image_data'] = image_map[r_idx]
            questions.append(q)

    else:
        # XLS - използваме xlrd
        BLACK_IDX = 8
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        for r in range(1, ws.nrows):
            q_val = ws.cell(r, 0).value
            if not q_val or str(q_val).strip() == '':
                continue
            q_text = str(q_val).strip()
            options = []
            opt_idx = 0

            for c in range(1, ws.ncols):
                cell = ws.cell(r, c)
                if not cell.value or str(cell.value).strip() == '':
                    continue
                text = str(cell.value).strip()
                xf_idx = ws.cell_xf_index(r, c)
                xf = wb.xf_list[xf_idx]
                font = wb.font_list[xf.font_index]
                is_correct = (font.colour_index != BLACK_IDX)
                options.append({
                    'letter': OPT_LETTERS[opt_idx] if opt_idx < len(OPT_LETTERS) else 'x',
                    'text': text,
                    'isCorrect': is_correct
                })
                opt_idx += 1

            if options and not any(o['isCorrect'] for o in options):
                options[0]['isCorrect'] = True

            questions.append({'id': len(questions) + 1, 'question': q_text, 'options': options})

    return questions

from app.utils.images import inject_images, save_test_images, delete_test_images


@admin.route('/tests/force-upload', methods=['POST'])
@admin_required
def force_upload():
    """Качва тест използвайки вече парснатите данни от сесията"""
    pending_file = session.get('pending_upload_file')
    if pending_file and __import__('os').path.exists(pending_file):
        with open(pending_file) as _pf:
            pending = __import__('json').load(_pf)
    else:
        pending = session.get('pending_upload')
    if not pending:
        return jsonify({'error': 'Няма данни за качване'}), 400
    
    new_title = request.json.get('title', pending['title'])
    
    test = Test(
        title=new_title,
        category=pending['category'],
        level=pending['level'],
        questions_json=pending['questions_json'],
        question_count=pending['question_count'],
        is_demo=False
    )
    db.session.add(test)
    db.session.flush()
    
    # Запази снимките — trайно в базата (или R2), не на диска на контейнера
    if pending.get('images_b64'):
        import base64
        images_to_save = [
            (qid, (base64.b64decode(b64data), fmt))
            for qid, b64data, fmt in pending['images_b64']
        ]
        print(f"FORCE_UPLOAD: Decoding {len(images_to_save)} pending images for test {test.id}")
        save_test_images(test.id, images_to_save)
    else:
        print("FORCE_UPLOAD: Няма pending снимки за този upload (нормално, ако тестът няма снимки)")

    db.session.commit()
    session.pop('pending_upload', None)
    return jsonify({'success': True, 'title': new_title, 'total': pending['question_count']})

# toggle_demo route removed - use /admin/demo/toggle/<id>

@admin.route('/tests/upload', methods=['POST'])
@admin_required
def upload_test():
    file = request.files.get('file')
    title = request.form.get('title', '').strip()
    category = request.form.get('category', 'deck')
    level = request.form.get('level', 'Operational Level')

    if not file:
        return jsonify({'error': 'Няма файл'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(tempfile.gettempdir(), filename)
    file.save(filepath)

    try:
        print(f"UPLOAD: Starting parse of {filename}, size={os.path.getsize(filepath)} bytes")
        questions = parse_xls_colors(filepath)
        print(f"UPLOAD: Parsed {len(questions)} questions")
        with_img = sum(1 for q in questions if q.get('has_image'))
        print(f"UPLOAD: Questions with images: {with_img}")
        final_title = title if title else filename.replace('.xls', '').replace('.xlsx', '')
        
        # Провери за дублиращо се заглавие
        existing = Test.query.filter_by(title=final_title).first()
        if existing:
            force = request.form.get('force', 'false')
            if force != 'true':
                # Запази парснатите данни в сесията за по-късно
                import pickle, base64
                # Запазваме в /tmp вместо в сесията (cookie limit)
                _pending_data = {
                    'questions_json': __import__('json').dumps(
                        [{k: v for k, v in q.items() if k != '_image_data'} for q in questions],
                        ensure_ascii=False
                    ),
                    'question_count': len(questions),
                    'category': category,
                    'level': level,
                    'title': final_title,
                    'images': [(q['id'], q['_image_data']) for q in questions if '_image_data' in q]
                }
                _pending_file = f'/tmp/pending_upload_{session.get("user_id","admin")}.json'
                with open(_pending_file, 'w') as _pf:
                    # ВАЖНО: images СА включени тук (base64), не изключени —
                    # преди тази поправка се губеха мълчаливо при force upload
                    # на тест със същото заглавие (никога не стигаха до
                    # save_test_images(), нямаше никаква грешка в логовете).
                    _pending_images_b64 = [
                        [qid, base64.b64encode(img_bytes).decode('ascii'), fmt]
                        for qid, (img_bytes, fmt) in _pending_data['images']
                    ]
                    __import__('json').dump({
                        **{k: v for k, v in _pending_data.items() if k != 'images'},
                        'images_b64': _pending_images_b64,
                    }, _pf)
                session['pending_upload_file'] = _pending_file
                session['pending_upload'] = {
                    'title': final_title,
                    'category': category,
                    'level': level,
                    'question_count': len(questions)
                }
                os.remove(filepath)
                return jsonify({'duplicate': True, 'title': final_title})
            else:
                # Намери следващия свободен индекс
                idx = 1
                while Test.query.filter_by(title=f"{final_title} ({idx})").first():
                    idx += 1
                final_title = f"{final_title} ({idx})" 

        # Извади снимките преди да запишем JSON
        images_to_save = []
        for q in questions:
            if '_image_data' in q:
                images_to_save.append((q['id'], q.pop('_image_data')))

        test = Test(
            title=final_title,
            category=category,
            level=level,
            questions_json=json.dumps(questions, ensure_ascii=False),
            question_count=len(questions),
            is_demo=False
        )
        db.session.add(test)
        db.session.flush()
        test_id_for_images = test.id
        db.session.commit()
        os.remove(filepath)

        # Запази снимките — trайно в базата, не на диска на контейнера
        if images_to_save:
            save_test_images(test_id_for_images, images_to_save)

        return jsonify({'success': True, 'total': len(questions), 'title': final_title})
    except Exception as e:
        try: os.remove(filepath)
        except: pass
        return jsonify({'error': str(e)}), 500

@admin.route('/tests/<int:test_id>/edit')
@admin_required
def edit_test(test_id):
    test = Test.query.get_or_404(test_id)
    questions = test.get_questions()
    questions = inject_images(test_id, questions)
    return render_template('admin/edit_test.html', test=test, questions=questions)

@admin.route('/tests/<int:test_id>/update-info', methods=['POST'])
@admin_required
def update_test_info(test_id):
    test = Test.query.get_or_404(test_id)
    data = request.json
    test.title = data.get('title', test.title)
    test.level = data.get('level', test.level)
    db.session.commit()
    return jsonify({'success': True})

@admin.route('/tests/<int:test_id>/delete', methods=['POST'])
@admin_required
def delete_test(test_id):
    test = Test.query.get_or_404(test_id)
    # Изтрий резултатите
    TestResult.query.filter_by(test_id=test_id).delete()
    db.session.delete(test)
    # Изтрий снимките от базата
    delete_test_images(test_id)
    db.session.commit()
    return jsonify({'success': True})

@admin.route('/tests/<int:test_id>/questions')
@admin_required
def get_test_questions(test_id):
    test = Test.query.get_or_404(test_id)
    return jsonify({'questions': test.get_questions(), 'title': test.title})

@admin.route('/tests/<int:test_id>/questions', methods=['POST'])
@admin_required
def save_test_questions(test_id):
    try:
        test = Test.query.get_or_404(test_id)
        questions = request.json.get('questions', [])

        # Запази has_image флага от оригиналните въпроси
        original = {str(q['id']): q for q in test.get_questions()}
        
        for q in questions:
            # Възстанови has_image от оригинала
            orig = original.get(str(q['id']))
            if orig and orig.get('has_image'):
                q['has_image'] = True

            # Гарантира само ЕДИН верен отговор
            correct_found = False
            for opt in q.get('options', []):
                if opt.get('isCorrect') and not correct_found:
                    correct_found = True
                elif opt.get('isCorrect') and correct_found:
                    opt['isCorrect'] = False
            if not correct_found and q.get('options'):
                q['options'][0]['isCorrect'] = True

        test.questions_json = json.dumps(questions, ensure_ascii=False)
        test.question_count = len(questions)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        import traceback
        print("SAVE QUESTIONS ERROR:", traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@admin.route('/users')
@admin_required
def admin_users():
    from app.models.gold_grant import GoldGrant
    now = datetime.utcnow()
    search_q = (request.args.get('q') or '').strip()
    users_query = User.query.filter_by(is_admin=False)
    if search_q:
        users_query = users_query.filter(
            db.or_(
                User.email.ilike(f'%{search_q}%'),
                db.cast(User.id, db.String).ilike(f'%{search_q}%'),
            )
        )
    users = users_query.order_by(User.created_at.desc()).all()

    user_ids = [u.id for u in users]
    grants_by_user = {}
    if user_ids:
        for g in GoldGrant.query.filter(GoldGrant.user_id.in_(user_ids), GoldGrant.expires_at > now).all():
            grants_by_user.setdefault(g.user_id, []).append(g)

    # Всеки текущо ВАЛИДЕН план/grant за потребителя — не user.plan (единично поле,
    # което не отразява, че може да има няколко активни Gold grant-а едновременно).
    plan_labels = {}
    for u in users:
        labels = []
        if u.plan in ('basic', 'plus') and u.plan_expires_at and u.plan_expires_at > now:
            labels.append(u.plan.upper())
        for g in grants_by_user.get(u.id, []):
            dept_short = (g.department or '?')[:4].capitalize()
            level_short = (g.level or '').split()[0][:3].upper() if g.level else ''
            labels.append(f"GOLD·{dept_short}{'/' + level_short if level_short else ''}")
        plan_labels[u.id] = labels or ['FREE']

    return render_template('admin/users.html', users=users, now=now, plan_labels=plan_labels, search_q=search_q)


@admin.route('/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        return jsonify({'success': False, 'message': 'Cannot delete admin'})
    db.session.delete(user)
    db.session.commit()
    return jsonify({'success': True})

@admin.route('/debug/plan-status')
@admin_required
def debug_plan_status():
    """
    Суровата истина за акаунт — без изчисления, без предположения.
    Използване: /admin/debug/plan-status?email=bumnazaloga3@abv.bg
    """
    from app.models.gold_grant import GoldGrant
    email = (request.args.get('email') or '').strip().lower()
    if not email:
        return jsonify({'error': 'Добави ?email=... в URL-a'}), 400

    user = User.query.filter(db.func.lower(User.email) == email).first()
    if not user:
        return jsonify({'error': f'Няма потребител с имейл {email}'}), 404

    now = datetime.utcnow()
    all_grants = GoldGrant.query.filter_by(user_id=user.id).order_by(GoldGrant.activated_at.desc()).all()

    return jsonify({
        'user_id': user.id,
        'email': user.email,
        'server_time_now': now.isoformat(),
        'RAW_DB_FIELDS': {
            'plan': user.plan,
            'is_active': user.is_active,
            'plan_activated_at': user.plan_activated_at.isoformat() if user.plan_activated_at else None,
            'plan_expires_at': user.plan_expires_at.isoformat() if user.plan_expires_at else None,
            'library_test_id': user.library_test_id,
            'tests_used': user.tests_used,
        },
        'COMPUTED_REAL_STATUS': {
            'has_active_plan': user.has_active_plan(),
            'effective_plan_label': user.effective_plan_label(),
            'effective_days_left': user.effective_days_left(),
        },
        'ALL_GOLD_GRANTS_IN_DB': [
            {
                'id': g.id,
                'promo_code': g.promo_code,
                'department': g.department,
                'level': g.level,
                'test_ids': g.test_id_list(),
                'quota': g.quota,
                'tests_used': g.tests_used,
                'activated_at': g.activated_at.isoformat() if g.activated_at else None,
                'expires_at': g.expires_at.isoformat() if g.expires_at else None,
                'IS_CURRENTLY_ACTIVE': g.expires_at > now if g.expires_at else False,
            }
            for g in all_grants
        ],
    })


@admin.route('/users/<int:user_id>')
@admin_required
def admin_user_detail(user_id):
    user = User.query.get_or_404(user_id)
    results = TestResult.query.filter_by(user_id=user_id).order_by(TestResult.taken_at.desc()).all()
    return render_template('admin/user_detail.html', user=user, results=results)

@admin.route('/users/<int:user_id>/billing')
@admin_required
def admin_user_billing(user_id):
    """
    Пълната billing история на потребителя (всички Basic/Plus/Gold покупки
    - активни И вече изтекли/използвани), за попъпа "Account" в admin/users.
    Същите данни, каквито потребителят вижда в собствения си Billing/Usage
    таб (grant.plan, кодa, activated_at/expires_at), но БЕЗ филтъра "само
    активните" - тук админът трябва да види ЦЯЛАТА история, включително
    колко пъти е ползвал платени абонаменти по-рано.
    """
    from app.models.plan_grant import PlanGrant
    from app.models.gold_grant import GoldGrant
    from app.utils.codes import get_or_create_subscription_code
    user = User.query.get_or_404(user_id)
    now = datetime.utcnow()

    cards = []
    all_plan_grants = PlanGrant.query.filter_by(user_id=user_id).order_by(PlanGrant.activated_at.asc()).all()
    for g in all_plan_grants:
        cards.append({
            'plan': g.plan.capitalize(),
            'code': get_or_create_subscription_code('plan', g.id),
            'activated_at': g.activated_at.strftime('%d.%m.%Y %H:%M') if g.activated_at else '—',
            'expires_at': g.expires_at.strftime('%d.%m.%Y %H:%M') if g.expires_at else '—',
            'status': 'Active' if g.expires_at and g.expires_at > now else 'Expired',
            '_sort_key': g.activated_at or datetime.min,
        })

    all_gold_grants = GoldGrant.query.filter_by(user_id=user_id).order_by(GoldGrant.activated_at.asc()).all()
    for g in all_gold_grants:
        cards.append({
            'plan': 'Gold',
            'code': get_or_create_subscription_code('gold', g.id),
            'activated_at': g.activated_at.strftime('%d.%m.%Y %H:%M') if g.activated_at else '—',
            'expires_at': g.expires_at.strftime('%d.%m.%Y %H:%M') if g.expires_at else '—',
            'status': 'Active' if g.expires_at and g.expires_at > now else 'Expired',
            '_sort_key': g.activated_at or datetime.min,
        })

    # Free-план сесии (library избор) - от FreeSession историята, СЪЩИЯ
    # формат като Basic/Plus/Gold картите по-горе, за да се вижда Free в
    # Usage/Billing попъпа на админа по абсолютно същия начин.
    from app.models.free_session import FreeSession
    free_cards = []
    all_free_sessions = FreeSession.query.filter_by(user_id=user_id).order_by(FreeSession.activated_at.asc()).all()
    for s in all_free_sessions:
        free_cards.append({
            'plan': 'Free',
            'code': f"{s.test.title[:22]}" if s.test else '—',
            'activated_at': s.activated_at.strftime('%d.%m.%Y %H:%M') if s.activated_at else '—',
            'expires_at': s.expires_at.strftime('%d.%m.%Y %H:%M') if s.expires_at else '—',
            'status': 'Active' if s.expires_at and s.expires_at > now else 'Expired',
            '_sort_key': s.activated_at or datetime.min,
        })

    all_cards_merged = cards + free_cards
    all_cards_merged.sort(key=lambda c: c['_sort_key'], reverse=True)
    for c in all_cards_merged:
        del c['_sort_key']

    return jsonify({
        'email': user.email,
        'server_time_utc': now.strftime('%Y-%m-%d %H:%M:%S'),
        'total_purchases': len(all_cards_merged),
        'cards': all_cards_merged,
    })

@admin.route('/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.email_verified = not user.email_verified
    db.session.commit()
    return jsonify({'success': True, 'email_verified': user.email_verified})

@admin.route('/promos')
@admin_required
def admin_promos():
    from app.models.payment import Payment
    now = datetime.utcnow()
    from app.models.gold_grant import GoldGrant
    promos = PromoCode.query.order_by(PromoCode.created_at.desc()).all()

    # payment date по stripe_payment_intent (за Gold кодове); fallback = created_at (ръчно създадени кодове)
    intents = [p.stripe_payment_intent for p in promos if p.stripe_payment_intent]
    payments_by_intent = {}
    if intents:
        for pay in Payment.query.filter(Payment.stripe_payment_intent.in_(intents)).all():
            payments_by_intent[pay.stripe_payment_intent] = pay

    # Grant-ове по promo код — реалният срок (спазва TESTING_MODE), не хардкоднати 30 дни
    grants_by_code = {}
    used_codes = [p.code for p in promos if p.is_used]
    if used_codes:
        for g in GoldGrant.query.filter(GoldGrant.promo_code.in_(used_codes)).all():
            grants_by_code[g.promo_code] = g

    rows = []
    for p in promos:
        payment = payments_by_intent.get(p.stripe_payment_intent)
        payment_date = payment.paid_at if payment else p.created_at

        if not p.is_used:
            status = 'expired' if (p.expires_at and p.expires_at < now) else 'stand-by'
        else:
            grant = grants_by_code.get(p.code)
            if grant:
                status = 'active' if grant.expires_at > now else 'used'
            else:
                # легаси код, активиран преди GoldGrant модела — няма грант запис.
                # Ползваме текущата конфигурация (спазва TESTING_MODE), не хардкоднати 30 дни.
                from app.services.plans import PLANS as _PLANS
                legacy_days = _PLANS['gold'].get('valid_days_per_code', 30)
                status = 'active' if (p.activated_at and (now - p.activated_at).days < legacy_days) else 'used'

        if p.is_used and not grants_by_code.get(p.code) and p.activated_at:
            from app.services.plans import PLANS as _PLANS2
            _legacy_days = _PLANS2['gold'].get('valid_days_per_code', 30)
            legacy_valid_until = p.activated_at + timedelta(days=_legacy_days)
        else:
            legacy_valid_until = None

        grant = grants_by_code.get(p.code)
        rows.append({
            'kind': 'gold', 'promo': p, 'code': p.code,
            'client_name': p.client_name, 'used_by': p.used_by, 'plan_label': 'Gold',
            'payment_date': payment_date, 'status': status,
            'valid_until': (grant.expires_at if p.is_used and grant else (legacy_valid_until or p.expires_at)),
            'seq_number': grant.id if grant else None,
        })

    # Basic/Plus плащания — нямат промокод (директна активация), обединяваме в същия списък.
    # Всяко плащане е свой собствен, автономен период на достъп — статусът му се смята
    # от собствения му прозорец (paid_at + план дни), а НЕ от това какъв е user.plan сега.
    from app.services.plans import get_plan_config
    from app.models.plan_grant import PlanGrant
    basic_plus_payments = Payment.query.filter(Payment.plan.in_(['basic', 'plus'])).all()
    for pay in basic_plus_payments:
        u = User.query.get(pay.user_id)
        if not u:
            continue
        cfg = get_plan_config(pay.plan) or {}
        days = cfg.get('days', 0)
        pay_expires = pay.paid_at + timedelta(days=days) if pay.paid_at and days else None
        bp_status = 'active' if (pay_expires and pay_expires > now) else 'used'

        grant = PlanGrant.query.filter_by(payment_id=pay.id).first()
        from app.utils.codes import get_or_create_subscription_code
        # Същият читаем BG код, ползван вече в Billing/Usage - не суровия
        # PlanGrant.id (само вътрешен database номер, безсмислен за админа
        # без контекст, а и лесен за объркване с GoldGrant.id при съвпадащ номер).
        unique_ref = get_or_create_subscription_code('plan', grant.id) if grant else None

        rows.append({
            'kind': pay.plan, 'promo': None, 'code': unique_ref,
            'client_name': u.email, 'used_by': u.email, 'plan_label': pay.plan.capitalize(),
            'payment_date': pay.paid_at, 'status': bp_status,
            'valid_until': pay_expires,
            'seq_number': grant.id if grant else None,
        })

    rows.sort(key=lambda r: r['payment_date'] or datetime.min, reverse=True)

    # Статистиките (Active/Used/Total) отразяват ВИНАГИ пълния набор от
    # данни, независимо от търсенето - търсенето филтрира само редовете в
    # самата таблица, не обобщените бройки горе.
    active = sum(1 for r in rows if r['status'] == 'active')
    used = sum(1 for r in rows if r['status'] in ('used', 'expired'))
    total_count = len(rows)

    # Търсене по email на клиента, по BG кода, или по суровия пореден номер
    # (seq_number) - същия UX паттерн като в admin/users.html (закръглена
    # кутийка с лупа вдясно).
    search_q = (request.args.get('q') or '').strip()
    if search_q:
        q_lower = search_q.lower()
        rows = [
            r for r in rows
            if q_lower in (r['client_name'] or '').lower()
            or q_lower in (r.get('used_by') or '').lower()
            or q_lower in (r['code'] or '').lower()
            or (r['seq_number'] is not None and q_lower in str(r['seq_number']))
        ]

    return render_template('admin/promos.html', rows=rows, promos=promos, active=active, used=used, total_count=total_count, search_q=search_q)

@admin.route('/promos/create', methods=['POST'])
@admin_required
def create_promo():
    client = request.form.get('client_name', '').strip()
    access_type = request.form.get('access_type', 'Регулярни тестове')
    price = float(request.form.get('price', 0) or 0)
    code = generate_promo_code()

    promo = PromoCode(code=code, client_name=client, access_type=access_type, price=price)
    db.session.add(promo)
    db.session.commit()
    return jsonify({'success': True, 'code': code})

@admin.route('/promos/<int:promo_id>/delete', methods=['POST'])
@admin_required
def delete_promo(promo_id):
    from app.models.gold_grant import GoldGrant
    promo = PromoCode.query.get_or_404(promo_id)

    # Изтриването на кода трябва реално да отнеме достъпа — иначе GoldGrant остава
    # жив в отделна таблица, независимо от промокода.
    affected_user = None
    if promo.used_by:
        affected_user = User.query.filter_by(email=promo.used_by).first()

    GoldGrant.query.filter_by(promo_code=promo.code).delete(synchronize_session=False)
    db.session.delete(promo)
    db.session.commit()

    # Синхронизираме плана на потребителя ВЕДНАГА — не да чака следваща проверка
    if affected_user:
        _sync_user_plan_after_revoke(affected_user)

    return jsonify({'success': True})

@admin.route('/promos/bulk-delete', methods=['POST'])
@admin_required
def bulk_delete_promos():
    from app.models.gold_grant import GoldGrant
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    ids = [int(i) for i in ids if str(i).isdigit()]
    if not ids:
        return jsonify({'success': False, 'message': 'No codes selected'}), 400

    promos = PromoCode.query.filter(PromoCode.id.in_(ids)).all()
    codes = [p.code for p in promos]
    affected_emails = {p.used_by for p in promos if p.used_by}

    if codes:
        GoldGrant.query.filter(GoldGrant.promo_code.in_(codes)).delete(synchronize_session=False)

    deleted = PromoCode.query.filter(PromoCode.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()

    # Синхронизираме плановете на всички засегнати потребители веднага
    for email in affected_emails:
        u = User.query.filter_by(email=email).first()
        if u:
            _sync_user_plan_after_revoke(u)

    return jsonify({'success': True, 'deleted': deleted})


def _sync_user_plan_after_revoke(user):
    """
    След премахване на GoldGrant — веднага обновява legacy полетата на потребителя
    (user.plan / is_active / plan_expires_at), ако вече няма никакъв валиден план.
    Иначе стар код, четящ директно тези полета, ще показва грешни данни до следваща
    случайна проверка.
    """
    if not user.has_active_plan():
        if user.plan == 'gold':
            user.plan = 'free'
            user.is_active = False
            user.plan_expires_at = None
            user.plan_activated_at = None
        db.session.commit()

@admin.route('/results/<int:result_id>')
@admin_required
def admin_result_detail(result_id):
    result = TestResult.query.get_or_404(result_id)
    test = Test.query.get(result.test_id)
    user = User.query.get(result.user_id)
    all_questions = test.get_questions()
    answers = json.loads(result.answers_json)

    # Ако имаме записани ID-та — показвай само тях
    try:
        q_ids = json.loads(result.question_ids_json or '[]')
    except:
        q_ids = []

    if q_ids:
        qid_set = set(str(q) for q in q_ids)
        questions = [q for q in all_questions if str(q['id']) in qid_set]
    else:
        answered_ids = set(answers.keys())
        questions = [q for q in all_questions if str(q['id']) in answered_ids] or all_questions

    # Зареди снимките
    questions = inject_images(result.test_id, questions)

    # Форматирай времето
    duration = result.duration or 0
    duration_str = f"{duration // 60:02d}:{duration % 60:02d}"

    # Тип на теста
    type_labels = {'test': 'Обикновен Тест', 'mix': 'Микс', 'simulator': 'Симулатор', 'mistakes': 'Грешки'}
    type_label = type_labels.get(result.test_type or 'test', 'Тест')

    return render_template('admin/result_detail.html',
        result=result, test=test, user=user,
        questions=questions, answers=answers,
        duration_str=duration_str, type_label=type_label)

@admin.route('/results/<int:result_id>/delete', methods=['POST'])
@admin_required
def delete_result(result_id):
    result = TestResult.query.get_or_404(result_id)
    db.session.delete(result)
    db.session.commit()
    return jsonify({'success': True})

@admin.route('/results/cleanup', methods=['POST'])
@admin_required
def cleanup_results():
    """Изтрива резултати по-стари от X дни"""
    days = int(request.json.get('days', 30))
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=days)
    old_results = TestResult.query.filter(TestResult.taken_at < cutoff).all()
    count = len(old_results)
    for r in old_results:
        db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True, 'deleted': count})

@admin.route('/results/cleanup-expired', methods=['POST'])
@admin_required
def cleanup_expired_results():
    """Изтрива резултати, чиито конкретен grant (по това време) вече е изтекъл — не по текущия общ статус на потребителя"""
    from app.models.gold_grant import GoldGrant
    from app.models.plan_grant import PlanGrant
    now = datetime.utcnow()
    all_results = TestResult.query.all()
    to_delete = []

    grants_cache = {}
    for r in all_results:
        if r.user_id not in grants_cache:
            grants_cache[r.user_id] = {
                'gold': GoldGrant.query.filter_by(user_id=r.user_id).all(),
                'plan': PlanGrant.query.filter_by(user_id=r.user_id).all(),
            }
        cache = grants_cache[r.user_id]

        is_active = False
        matched = False
        for g in cache['gold']:
            if r.test_id in g.test_id_list() and g.activated_at and g.activated_at <= r.taken_at:
                is_active = g.expires_at > now
                matched = True
                break
        if not matched:
            for g in cache['plan']:
                if g.library_test_id == r.test_id and g.activated_at and g.activated_at <= r.taken_at:
                    is_active = g.expires_at > now
                    matched = True
                    break

        if not is_active:
            to_delete.append(r)

    count = len(to_delete)
    for r in to_delete:
        db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True, 'deleted': count})


@admin.route('/signals')
@admin_required
def admin_signals():
    signals = Signal.query.order_by(Signal.created_at.desc()).all()
    open_count = Signal.query.filter_by(status='open').count()
    return render_template('admin/signals.html', signals=signals, open_count=open_count)

@admin.route('/signals/<int:signal_id>/resolve', methods=['POST'])
@admin_required
def resolve_signal(signal_id):
    signal = Signal.query.get_or_404(signal_id)
    signal.status = 'resolved'
    db.session.commit()
    return jsonify({'success': True})

# ============================================================
#  ИНИЦИАЛИЗАЦИЯ
# ============================================================

from app.utils.codes import alternating_code, subscription_code, result_public_code, get_or_create_subscription_code
from app.utils.grants import find_result_grant as _find_result_grant
from app.utils.grants import auto_delete_expired_results as _auto_delete_expired_results


@admin.route('')
@admin_required
def admin_dashboard():
    from app.services.stats import get_admin_stats
    from app.models.result import TestResult
    stats = get_admin_stats()
    admin_user = User.query.filter_by(is_admin=True).first()
    now = datetime.utcnow()

    # Опортюнистично автоматично почистване — 30 дни grace период след изтичане
    auto_deleted = _auto_delete_expired_results()

    # Търсене в историята — по имейл на регистрация или по ID/display_id на резултата
    search_q = (request.args.get('q') or '').strip()
    results_query = (TestResult.query
                      .options(db.joinedload(TestResult.user), db.joinedload(TestResult.test))
                      .order_by(TestResult.taken_at.desc()))
    if search_q:
        results_query = results_query.join(User, TestResult.user_id == User.id).filter(
            db.or_(
                User.email.ilike(f'%{search_q}%'),
                db.cast(TestResult.id, db.String).ilike(f'%{search_q}%'),
            )
        )
        recent_results = results_query.limit(50).all()
    else:
        recent_results = results_query.limit(10).all()

    # Статус на плана — ПО РЕЗУЛТАТ, не по потребител! Намираме КОНКРЕТНИЯ grant,
    # който е покривал точно ТОЗИ тест по времето на решаването му, и проверяваме
    # дали ИМЕННО ТОЗИ grant все още е активен — не дали потребителят има ДРУГ,
    # несвързан, по-нов план в момента (иначе стар изтекъл резултат лъжливо
    # показва "Active" само защото user-ът е активирал нещо ново оттогава).
    plan_status_by_result_id = {}
    public_code_by_result_id = {}
    from app.models.gold_grant import GoldGrant
    from app.models.plan_grant import PlanGrant
    _unique_uids = list({r.user_id for r in recent_results})
    _all_gold = GoldGrant.query.filter(GoldGrant.user_id.in_(_unique_uids)).all() if _unique_uids else []
    _all_plan = PlanGrant.query.filter(PlanGrant.user_id.in_(_unique_uids)).all() if _unique_uids else []
    gold_cache, plan_cache = {}, {}
    for _uid in _unique_uids:
        gold_cache[_uid] = [g for g in _all_gold if g.user_id == _uid]
        plan_cache[_uid] = [g for g in _all_plan if g.user_id == _uid]
    for r in recent_results:
        status, grant = _find_result_grant(r, now, gold_cache, plan_cache)
        plan_status_by_result_id[r.id] = status

        if grant:
            grant_test_ids = grant.test_id_list() if hasattr(grant, 'test_id_list') else [grant.library_test_id]
            seq = (TestResult.query
                   .filter(TestResult.user_id == r.user_id,
                           TestResult.test_id.in_(grant_test_ids),
                           TestResult.taken_at >= grant.activated_at,
                           TestResult.taken_at <= r.taken_at)
                   .count())
            _grant_type = 'gold' if hasattr(grant, 'test_id_list') else 'plan'
            # ПОПРАВКА (същия бъг като user-ската история, вижте dashboard.py):
            # за Gold ползваме РЕАЛНИЯ активиран код (grant.promo_code), не
            # преизчислен нов от grant.id.
            if _grant_type == 'gold':
                _base_code = grant.promo_code or subscription_code(grant.id, grant_type='gold')
            else:
                _base_code = get_or_create_subscription_code('plan', grant.id)
            public_code_by_result_id[r.id] = f"{_base_code}{r.taken_at.strftime('%d%m%y')}-{seq:03d}"
        else:
            public_code_by_result_id[r.id] = None

    recent_signals = []
    return render_template('admin/dashboard.html',
        admin_user=admin_user,
        public_code_by_result_id=public_code_by_result_id,
        recent_results=recent_results,
        plan_status_by_result_id=plan_status_by_result_id,
        search_q=search_q,
        auto_deleted=auto_deleted,
        recent_signals=recent_signals,
        **stats)

@admin.route('/tests')
@admin_required
def admin_tests():
    deck_tests = Test.query.filter_by(category='deck').order_by(Test.created_at.desc()).all()
    engine_tests = Test.query.filter_by(category='engine').order_by(Test.created_at.desc()).all()
    admin_user = User.query.filter_by(is_admin=True).first()
    deck_q = sum(t.question_count for t in deck_tests)
    engine_q = sum(t.question_count for t in engine_tests)
    mistakes_ready = False
    demo_sessions = 0
    return render_template('admin/tests.html',
        deck_tests=deck_tests, engine_tests=engine_tests,
        deck_q=deck_q, engine_q=engine_q,
        mistakes_ready=mistakes_ready, admin_user=admin_user)

@admin.route('/demo')
@admin_required
def admin_demo():
    tests = Test.query.filter_by(is_demo=True).order_by(Test.created_at.desc()).all()
    deck_demo = sum(1 for t in tests if t.category == 'deck')
    engine_demo = sum(1 for t in tests if t.category == 'engine')
    demo_count = len(tests)
    admin_user = User.query.filter_by(is_admin=True).first()
    return render_template('admin/demo.html',
        tests=tests, deck_demo=deck_demo,
        engine_demo=engine_demo, demo_count=demo_count,
        admin_user=admin_user)

@admin.route('/demo/toggle/<int:test_id>', methods=['POST'])
@admin_required
def admin_demo_toggle(test_id):
    test = Test.query.get_or_404(test_id)
    test.is_demo = not test.is_demo
    db.session.commit()
    return jsonify({'success': True, 'is_demo': test.is_demo})

@admin.route('/tests/next-title')
@admin_required
def next_title():
    title = request.args.get('title', '').strip()
    if not title:
        return jsonify({'exists': False, 'title': title})
    existing = Test.query.filter_by(title=title).first()
    if not existing:
        return jsonify({'exists': False, 'title': title})
    counter = 2
    while True:
        new_title = f"{title} ({counter})"
        if not Test.query.filter_by(title=new_title).first():
            return jsonify({'exists': True, 'title': new_title})
        counter += 1

@admin.route('/support/start/<int:user_id>', methods=['POST'])
@admin_required
def admin_support_start(user_id):
    """
    Admin-ът стартира НОВ разговор с потребител, който още няма никакъв
    ticket - преди тази промяна нямаше начин admin да ИНИЦИИРА съобщение,
    само да отговаря на вече съществуващи, отворени от потребителя tickets.
    """
    from app.models.ticket import TicketMessage
    user = User.query.get_or_404(user_id)
    body = (request.get_json(silent=True) or {}).get('body', '').strip()
    if not body:
        return jsonify({'success': False, 'message': 'Empty message'}), 400

    ticket = Ticket(user_id=user.id, subject='Admin message', type='question', status='in_progress')
    db.session.add(ticket)
    db.session.flush()
    msg = TicketMessage(ticket_id=ticket.id, sender='admin', body=body, is_read=False)
    db.session.add(msg)
    db.session.commit()
    return jsonify({'success': True, 'ticket_id': ticket.id})

@admin.route('/support')
@admin_required
def admin_support():
    """
    ВАЖНО: темплейтът очаква {% set t = item.ticket %}{% set u = item.user %}
    и item.unread за всеки ред - преди тази поправка тук се подаваха голи
    Ticket обекти директно (tickets_query.all()), които нямат .ticket/.user
    атрибути -> Jinja UndefinedError -> 500 грешка на ВСЯКА заявка към тази
    страница, щом има поне 1 реален ticket в базата. Точно затова 'Message
    in Support Chat' изглеждаше 'несвързано' - страницата зад него беше
    напълно счупена.
    """
    from types import SimpleNamespace
    from app.models.ticket import TicketMessage

    filter_user_id = request.args.get('user_id', type=int)
    tickets_query = Ticket.query.order_by(Ticket.created_at.desc())
    if filter_user_id:
        tickets_query = tickets_query.filter_by(user_id=filter_user_id)
    raw_tickets = tickets_query.all()

    tickets = []
    for t in raw_tickets:
        unread = TicketMessage.query.filter_by(ticket_id=t.id, sender='user', is_read=False).count()
        u = User.query.get(t.user_id)
        tickets.append(SimpleNamespace(ticket=t, user=u, unread=unread))

    admin_user = User.query.filter_by(is_admin=True).first()
    filter_user = User.query.get(filter_user_id) if filter_user_id else None
    return render_template('admin/support.html', tickets=tickets, admin_user=admin_user, filter_user=filter_user)

@admin.route('/support/tickets')
@admin_required
def admin_support_tickets():
    from app.models.ticket import TicketMessage
    tickets = Ticket.query.order_by(Ticket.updated_at.desc()).all()
    result = []
    for t in tickets:
        unread = TicketMessage.query.filter_by(ticket_id=t.id, sender='user', is_read=False).count()
        user = User.query.get(t.user_id)
        result.append({
            'id': t.id, 'email': user.email if user else '', 'name': user.name if user else '',
            'type': t.type, 'status': t.status, 'unread': unread,
            'created_at': t.created_at.strftime('%d.%m %H:%M')
        })
    return jsonify(result)

@admin.route('/support/<int:ticket_id>/messages')
@admin_required
def admin_ticket_messages(ticket_id):
    from app.models.ticket import TicketMessage
    ticket = Ticket.query.get_or_404(ticket_id)
    user = User.query.get(ticket.user_id)
    messages = TicketMessage.query.filter_by(ticket_id=ticket_id).order_by(TicketMessage.created_at).all()
    TicketMessage.query.filter_by(ticket_id=ticket_id, sender='user', is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({
        'ticket': {'id': ticket.id, 'type': ticket.type, 'status': ticket.status},
        'user': {'email': user.email if user else '', 'name': user.name if user else ''},
        'messages': [{'id': m.id, 'body': m.body, 'sender': m.sender,
                      'created_at': m.created_at.strftime('%d.%m %H:%M')} for m in messages]
    })

@admin.route('/support/<int:ticket_id>/reply', methods=['POST'])
@admin_required
def admin_ticket_reply(ticket_id):
    from app.models.ticket import TicketMessage
    ticket = Ticket.query.get_or_404(ticket_id)
    body = request.form.get('body', '').strip()
    if not body:
        return jsonify({'success': False})
    msg = TicketMessage(ticket_id=ticket_id, sender='admin', body=body, is_read=False)
    ticket.status = 'in_progress'
    ticket.updated_at = datetime.utcnow()
    db.session.add(msg)
    db.session.commit()
    return jsonify({'success': True})

@admin.route('/support/<int:ticket_id>/close', methods=['POST'])
@admin_required
def admin_ticket_close(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    ticket.status = 'closed'
    ticket.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True})

@admin.route('/support/unread')
@admin_required
def admin_support_unread():
    from app.models.ticket import TicketMessage
    count = TicketMessage.query.filter_by(sender='user', is_read=False).count()
    return jsonify({'count': count})

@admin.route('/support/stats')
@admin_required
def admin_support_stats():
    pending = Ticket.query.filter(Ticket.status != 'closed').count()
    total = Ticket.query.count()
    return jsonify({'pending': pending, 'total': total})

@admin.route('/api/snapshots/<metric>')
@admin_required
def admin_snapshots(metric):
    from app.services.stats import get_admin_stats
    period = request.args.get('period', '1Y')
    snapshots = MonthlySnapshot.query.order_by(MonthlySnapshot.recorded_at).all()
    labels = [s.recorded_at.strftime('%b %Y') for s in snapshots]
    data = [getattr(s, metric, 0) or 0 for s in snapshots]
    return jsonify({'metric': metric, 'labels': labels, 'data': data})

@admin.route('/api/snapshots/record', methods=['POST'])
@admin_required
def admin_record_snapshot():
    from app.services.stats import record_monthly_snapshot
    record_monthly_snapshot()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# Еднократна поправка: акаунти, автоматично ъпгрейднати на Gold от стар бъг
# (webhook-ът задаваше user.plan='gold' директно на купувача, вместо да чака
# той сам да активира код през /activate). Засегнати: plan='gold' И
# gold_test_ids празно (никога не са минали през реалната активация).
# ---------------------------------------------------------------------------

def _find_gold_autobug_users():
    """Връща list от (user, proposed_plan, reason) за преглед преди поправка."""
    affected = User.query.filter(
        User.plan == 'gold',
        User.gold_test_ids.is_(None)
    ).all()

    results = []
    now = datetime.utcnow()
    for u in affected:
        # Търсим последно платено НЕ-gold плащане — то не е пипано от бъга
        last_other = (Payment.query
                      .filter(Payment.user_id == u.id, Payment.plan != 'gold')
                      .order_by(Payment.paid_at.desc())
                      .first())

        if last_other and u.plan_expires_at and u.plan_expires_at > now:
            proposed_plan = last_other.plan
            reason = f"Има валиден {last_other.plan} до {u.plan_expires_at.strftime('%d.%m.%Y')} (плащане от {last_other.paid_at.strftime('%d.%m.%Y')})"
        else:
            proposed_plan = 'free'
            reason = "Няма валидно предишно плащане с неизтекъл достъп — връщаме на free"

        results.append({'user': u, 'proposed_plan': proposed_plan, 'reason': reason})
    return results


@admin.route('/fix-gold-autobug')
@admin_required
def fix_gold_autobug_preview():
    """Dry-run — само показва какво ще се промени, нищо не пипа."""
    rows = _find_gold_autobug_users()
    return render_template('admin/fix_gold_autobug.html', rows=rows)


@admin.route('/fix-gold-autobug/apply', methods=['POST'])
@admin_required
def fix_gold_autobug_apply():
    rows = _find_gold_autobug_users()
    now = datetime.utcnow()
    fixed = 0

    for row in rows:
        u = row['user']
        if row['proposed_plan'] == 'free':
            u.plan = 'free'
            u.is_active = False
            u.plan_activated_at = None
            u.plan_expires_at = None
        else:
            u.plan = row['proposed_plan']
            u.is_active = True
        fixed += 1

    db.session.commit()
    flash(f'Поправени {fixed} акаунта, засегнати от Gold auto-upgrade бъга.', 'success')
    return redirect(url_for('admin.admin_promos'))


# ---------------------------------------------------------------------------
# Реклами (Free план + demo тестове) — показвани на всеки 5-ти въпрос
# ---------------------------------------------------------------------------

@admin.route('/ads')
@admin_required
def admin_ads():
    from app.models.ad import Ad
    ads = Ad.query.order_by(Ad.created_at.desc()).all()
    return render_template('admin/ads.html', ads=ads)


@admin.route('/ads/create', methods=['POST'])
@admin_required
def create_ad():
    from app.models.ad import Ad
    ad = Ad(
        title=request.form.get('title', '').strip(),
        image_url=request.form.get('image_url', '').strip() or None,
        link_url=request.form.get('link_url', '').strip() or None,
        body=request.form.get('body', '').strip() or None,
        is_active=True,
    )
    if not ad.title:
        return jsonify({'success': False, 'message': 'Заглавието е задължително.'}), 400
    db.session.add(ad)
    db.session.commit()
    return jsonify({'success': True, 'id': ad.id})


@admin.route('/ads/<int:ad_id>/toggle', methods=['POST'])
@admin_required
def toggle_ad(ad_id):
    from app.models.ad import Ad
    ad = Ad.query.get_or_404(ad_id)
    ad.is_active = not ad.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': ad.is_active})


@admin.route('/ads/<int:ad_id>/delete', methods=['POST'])
@admin_required
def delete_ad(ad_id):
    from app.models.ad import Ad
    ad = Ad.query.get_or_404(ad_id)
    db.session.delete(ad)
    db.session.commit()
    return jsonify({'success': True})
